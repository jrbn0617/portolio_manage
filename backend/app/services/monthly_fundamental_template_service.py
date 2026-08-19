"""월간 펀더멘털 DataGuide 요청 양식(WISEfn Add-in 포맷) 생성.

**원래는 `reference/monthly_data_template.xlsx`의 'sample' 시트를 복제해 채웠는데,
`reference/`가 gitignore 대상이라 다른 PC에는 그 파일이 없다.** 실제로 이 저장소를
받은 상태에서는 양식을 만들 수 없었다. 그래서 레이아웃을 코드로 옮겨 파일 의존을 없앴다
— 기존에 만들어 둔 요청 파일(reference/요청_*.xlsx)에서 셀 배치를 그대로 확인했다.

시트 레이아웃 (1-indexed)
  1행  Refresh / Last Update      8행  Code       (A+티커)
  3행  Time Series                9행  Name
  4행  Frequency / M / Ascending 10행  Item Code
  5행  Period(From) / Korean     11행  Unit
  6행  Period(To)                12행  Base Date
                                14행  D A T E / 항목명
                                15행~ 월말 날짜 (값은 비워서 보내고 채워서 돌려받는다)

이 배치는 `monthly_fundamental_bulk_service`가 읽는 위치와 같아야 한다 —
그쪽은 0-indexed로 CODE_ROW=7, DATA_START_ROW=14 를 본다. **시트명 접두어도
FIELD_METRIC_MAP과 일치해야** 업로드가 그 시트를 찾는다.
"""
import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.models.market_holiday import MarketHoliday
from app.services.instrument_rules import is_common_stock

# A1은 WISEfn/Quantiwise Add-in의 **Refresh 버튼**이다. 도형도 폼 컨트롤도 아니고
# 두 가지가 겹쳐 있다 (reference/백필요청1_가격_201401.xlsx를 풀어 확인):
#   ① 버튼처럼 보이게 하는 셀 서식 (흰 글씨 + 파란 배경)
#   ② 클릭을 받는 **하이퍼링크** — <hyperlink ref="A1" tooltip="Quantiwise7G"/>
# ②가 없으면 셀일 뿐이라 눌러도 아무 일도 일어나지 않는다. Add-in이 하이퍼링크 클릭
# 이벤트를 가로채는 방식이고, 링크 대상(target)은 없고 tooltip이 Add-in 식별자다.
# vmlDrawing은 전부 메모(Note)였고 버튼과 무관하다.
FONT_NAME = "맑은 고딕"
BUTTON_FILL = "FF5A7AB2"
BUTTON_FONT = "FFFFFFFF"
LABEL_COLOR = "FF9B9B9B"   # Code/Name/Item Code/Unit/Base Date 라벨
TITLE_COLOR = "FF404040"   # 'Time Series'
COL_WIDTH = 14.625         # 원본은 1~831열 전부 같은 폭이다
BUTTON_TOOLTIP = "Quantiwise7G"  # Add-in 식별자. 이 값으로 클릭을 잡는다
META_LAST_COL = 4          # Frequency~Period(To) 블록의 테두리는 D열까지만 그어진다
DATE_FMT = "mm-dd-yy"      # 원본이 쓰는 서식 그대로

# 메타 입력칸의 설명 메모 — 원본에 있는 것을 그대로 옮겼다. 양식을 손으로 고칠 때
# 어떤 값을 넣어야 하는지 알려주는 안내라 빠지면 쓰기 불편해진다.
CELL_NOTES = {
    "B4": "D - Daily\nW - Weekly\nM - Monthly\nQ - Quarterly\nY - Yearly",
    "C4": "A - Ascending\nD - Descending",
    "D4": "0 - Business\n1 - Previous\n2 - #N/A\n3 - Null",
    "C5": "K - Korean\nE - English",
    "D5": "0 - Non Sat,Sun\n1 - Sat\n2 - Sun\n3 - Sat,Sun",
    "B6": ("CPD     : 당일 (Calendar기준)\nCPD-1   : 전일 (Calendar기준)\n"
           "CPD-1TD : 직전 영업일\nCPW-1W  : 직전 영업일 기준 1주 전 주말(영업일 기준)\n"
           "CPM-1M  : 직전 영업일 기준 1달 전 월말(영업일 기준)"),
    "D6": "Account Download\n0 - Full\n1 - Point",
}

CODE_ROW, NAME_ROW, ITEM_CODE_ROW, UNIT_ROW, BASE_DATE_ROW = 8, 9, 10, 11, 12
HEADER_ROW, DATE_START_ROW = 14, 15
CHUNK_SIZE = 500  # WISEfn Add-in이 한 시트에서 받는 종목 수
DEFAULT_LOOKBACK_MONTHS = 6

# (시트명 접두어, Item Code, Unit, Base Date, 헤더 표기)
# 접두어는 monthly_fundamental_bulk_service.FIELD_METRIC_MAP과 반드시 일치해야 한다.
ITEM_SPECS = [
    ("유동비율", "S102060", "%", "", "유동비율"),
    ("EBITDA(TTM)", "M123005.M", "Local thou", "NR.FY1", "EBITDA(TTM)"),
    ("EBITDA(Fwd.12M)", "E123060.M", "Local thou", "", "EBITDA(Fwd.12M)"),
    ("EV EBITDA(Fwd.12M)", "E331060.M", "X", "", "EV/EBITDA(Fwd.12M)"),
]


def _shift_month(year: int, month: int, offset: int) -> tuple[int, int]:
    total = year * 12 + (month - 1) + offset
    return total // 12, total % 12 + 1


def _last_calendar_day(year: int, month: int) -> datetime.date:
    ny, nm = _shift_month(year, month, 1)
    return datetime.date(ny, nm, 1) - datetime.timedelta(days=1)


def holiday_coverage_through(db: Session) -> datetime.date | None:
    """market_holidays가 어디까지 채워져 있는지. 그 뒤 날짜는 휴장 여부를 알 수 없다."""
    return db.query(func.max(MarketHoliday.date)).scalar()


def month_end_business_days(db: Session, year: int, month: int, lookback: int) -> list[datetime.date]:
    """각 달의 **마지막 영업일**. 월 단위 데이터라 요청 기준일은 항상 월말이어야 한다.

    prices에서 그 달의 최대 날짜를 쓰면 안 된다 — 진행 중인 달에서는 '오늘'이 잡혀서
    요청일이 월말이 아니게 된다(예: 8월분을 8/18로 요청). 달력에서 말일부터 거꾸로
    걸어가며 주말과 market_holidays를 건너뛰는 방식으로 구한다.

    주의: market_holidays는 지나간 날짜만 채워져 있다(daily_update가 매일 적재).
    아직 오지 않은 달의 말일이 휴장일이면 그 사실을 알 수 없어 그대로 잡힌다 —
    holiday_coverage_through()로 확인할 수 있다."""
    months = [_shift_month(year, month, o) for o in range(-(lookback - 1), 1)]
    first = datetime.date(months[0][0], months[0][1], 1)
    last = _last_calendar_day(*months[-1])
    holidays = {r[0] for r in db.query(MarketHoliday.date)
                .filter(MarketHoliday.date.between(first, last)).all()}

    out = []
    for y, m in months:
        d = _last_calendar_day(y, m)
        while d.weekday() >= 5 or d in holidays:
            d -= datetime.timedelta(days=1)
        out.append(d)
    return out


def fetch_universe(db: Session, start: datetime.date, end: datetime.date) -> list[tuple[str, str]]:
    """[start, end] 구간에 코스피/코스닥에 **실제로 있었던** 보통주.

    두 조건을 따로 본다.
      - 시장 소속: index_memberships(KOSPI/KOSDAQ)에 이력이 있는가
      - 기간 내 존재: 그 구간에 일별 시세가 있는가

    왜 이렇게 나눴나 —
      · `instruments.market`으로는 못 거른다. 값이 'KOSPI'와 '유가증권시장'으로 뒤섞여
        있고 195건은 비어 있다.
      · index_memberships만으로도 안 된다. 스냅샷이 성기다(2015~2026에 32일치뿐,
        과거는 반년 간격). 구간 중간에 상장폐지된 종목이 스냅샷에 안 걸려 빠진다.
      · 시세는 매일 있으므로 '그 기간에 존재했는가'는 시세로 판정하는 게 정확하다.
        구간 이전에 폐지된 옛 종목은 시세가 없어 자연히 빠진다.

    우선주 판별은 **종목코드 끝자리**로 한다 — 이름 정규식은 오탐/누락이 있다
    (`연우`·`성우`가 보통주인데 걸리고, 숫자 없는 `삼성전자우` 형태는 놓친다)."""
    rows = db.execute(text("""
        SELECT i.ticker, i.name FROM instruments i
        WHERE i.asset_type = 'stock'
          AND EXISTS (SELECT 1 FROM index_memberships m
                      WHERE m.instrument_id = i.id AND m.index_name IN ('KOSPI', 'KOSDAQ'))
          AND EXISTS (SELECT 1 FROM prices p
                      WHERE p.instrument_id = i.id AND p.period = 'D'
                        AND p.date BETWEEN :start AND :end)
        ORDER BY i.ticker
    """), {"start": start, "end": end}).all()
    return [(t, n) for t, n in rows if is_common_stock(t)]


THIN = Side(style="thin")


def _style_sheet(ws, n_cols: int, n_dates: int):
    """원본 요청 파일의 서식을 그대로 입힌다 — 특히 A1(Refresh 버튼)."""
    ws.column_dimensions["A"].width = COL_WIDTH
    for i in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTH
    ws.row_dimensions[1].height = 15

    for row in ws.iter_rows(min_row=1, max_row=DATE_START_ROW + n_dates,
                            min_col=1, max_col=n_cols + 1):
        for c in row:
            c.font = Font(name=FONT_NAME, size=11)

    a1 = ws["A1"]
    a1.font = Font(name=FONT_NAME, size=9, color=BUTTON_FONT)
    a1.fill = PatternFill("solid", fgColor=BUTTON_FILL)
    a1.alignment = Alignment(horizontal="center", vertical="center")
    # 클릭을 받는 실체. target 없이 tooltip만 있는 하이퍼링크다.
    a1.hyperlink = Hyperlink(ref="A1", tooltip=BUTTON_TOOLTIP, display=a1.value)

    for addr, note in CELL_NOTES.items():
        ws[addr].comment = Comment(note, "Kiwoom", height=110, width=220)

    # A2의 0은 흰 글씨라 화면에서 보이지 않는다 (Add-in이 쓰는 플래그)
    ws["A2"].font = Font(name=FONT_NAME, size=11, color=BUTTON_FONT)
    ws["A3"].font = Font(name=FONT_NAME, size=11, bold=True, color=TITLE_COLOR)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    # 메타 블록(Frequency~Period(To))은 D열까지 좌측 정렬이다
    for r in (4, 5, 6):
        for c in range(1, META_LAST_COL + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="left", vertical="center")
    ws["B5"].number_format = DATE_FMT
    ws["B6"].number_format = "yyyy\\-mm\\-dd"

    for r in range(CODE_ROW, BASE_DATE_ROW + 2):
        cell = ws.cell(row=r, column=1)
        cell.font = Font(name=FONT_NAME, size=11, color=LABEL_COLOR)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{HEADER_ROW}"].font = Font(name=FONT_NAME, size=11)
    ws[f"A{HEADER_ROW}"].alignment = Alignment(horizontal="center", vertical="center")

    # 테두리 — 메타 블록(4·6행)은 D열까지, 데이터 블록(8·9·14행)은 마지막 종목열까지.
    for c in range(1, META_LAST_COL + 1):
        ws.cell(row=4, column=c).border = Border(top=THIN)
        ws.cell(row=6, column=c).border = Border(bottom=THIN)
    for c in range(1, n_cols + 2):
        ws.cell(row=CODE_ROW, column=c).border = Border(top=THIN)
        ws.cell(row=HEADER_ROW, column=c).border = Border(bottom=THIN)
        if c > 1:  # 원본에서 A9만 아래 테두리가 없다
            ws.cell(row=NAME_ROW, column=c).border = Border(bottom=THIN)


def _fill_sheet(ws, chunk, dates, item_code, unit, base_date, header):
    ws["A1"] = "     Refresh     "
    ws["B1"] = f"Last Update : {datetime.datetime.now():%Y-%m-%d %H:%M:%S}"
    ws["A2"] = 0
    ws["A3"] = "Time Series"
    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Frequency", "M", "Ascending", 0
    ws["A5"], ws["B5"], ws["C5"], ws["D5"] = "Period(From)", dates[0], "Korean", 0
    ws["A6"], ws["B6"], ws["D6"] = "Period(To)", dates[-1], 0
    ws[f"A{CODE_ROW}"] = "Code"
    ws[f"A{NAME_ROW}"] = "Name"
    ws[f"A{ITEM_CODE_ROW}"] = "Item Code"
    ws[f"A{UNIT_ROW}"] = "Unit"
    ws[f"A{BASE_DATE_ROW}"] = "Base Date"
    ws[f"A{HEADER_ROW}"] = "D A T E"

    for i, (ticker, name) in enumerate(chunk):
        col = get_column_letter(i + 2)
        ws[f"{col}{CODE_ROW}"] = f"A{ticker}"
        ws[f"{col}{NAME_ROW}"] = name
        ws[f"{col}{ITEM_CODE_ROW}"] = item_code
        ws[f"{col}{UNIT_ROW}"] = unit
        if base_date:
            ws[f"{col}{BASE_DATE_ROW}"] = base_date
        ws[f"{col}{HEADER_ROW}"] = header

    for i, d in enumerate(dates):
        cell = ws.cell(row=DATE_START_ROW + i, column=1, value=d)
        cell.number_format = DATE_FMT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _style_sheet(ws, len(chunk), len(dates))


def build_template(db: Session, year: int, month: int,
                   lookback: int = DEFAULT_LOOKBACK_MONTHS) -> openpyxl.Workbook:
    dates = month_end_business_days(db, year, month, lookback)
    # 유니버스는 요청 구간(첫 달 1일 ~ 마지막 달 말일) 기준으로 뽑는다 — 그 구간에
    # 없던 옛 종목을 요청에 넣으면 응답이 통째로 빈 열로 돌아온다.
    universe = fetch_universe(db, datetime.date(dates[0].year, dates[0].month, 1), dates[-1])
    if not universe:
        raise RuntimeError("요청 구간에 코스피/코스닥 종목이 없습니다.")
    chunks = [universe[i:i + CHUNK_SIZE] for i in range(0, len(universe), CHUNK_SIZE)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for prefix, item_code, unit, base_date, header in ITEM_SPECS:
        for n, chunk in enumerate(chunks, start=1):
            _fill_sheet(wb.create_sheet(f"{prefix}{n}"), chunk, dates, item_code, unit, base_date, header)
    return wb
