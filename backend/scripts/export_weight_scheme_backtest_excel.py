"""실험#10/#11(가중치 방식 + 섹터캡 변형) 백테스트 결과를
reference/(별첨 2) 백테스팅자료_회사명(예시).xlsx 포맷으로 출력한다.

최종 채택안(실험#9: EBITDA PEG 스크리닝+모멘텀랭킹, 월중손절 -10%(익일시가체결),
시장레짐필터(200일선, 약세시 50%))을 기반으로 섹터 관련 제약을 바꿔가며 비교한다.
--max-per-sector로 카운트 기반 제한(기본값 None=제한없음, 실험#9와 같이 2를 주면
"섹터당 최대 2종목" 유지)을, --group-field/--max-group-weight로 합산비중 기반 그룹캡을
독립적으로 켜고 끌 수 있다:
  실험#10: --max-per-sector 없음(제한 제거) + equal 또는 free_float+group-field(3종) — 기각
  실험#11: --max-per-sector 2(유지) + free_float+group-field=sector,max-group-weight=0.50
           (기존 섹터당2종목 캡에 동일섹터 합산 50%캡을 추가로 병행)
2~4는 기존 종목당 30% 상한(--max-weight)도 그룹캡과 함께 병행 적용한다
(app/services/backtest_service.cap_weights_with_groups).

사용법: python scripts/export_weight_scheme_backtest_excel.py [--index KOSDAQ150]
        [--top-n 20] [--max-per-sector 2|없음] [--weighting equal|free_float]
        [--group-field none|krx_sector|sector|industry]
        [--max-weight 0.30] [--max-group-weight 0.50]
        [--stop-loss-pct 0.10] [--stop-loss-execution next_open]
        [--ma-window-days 200] [--bear-exposure 0.5]
        [--start 2020-01-01] [--end 2026-07-31] [--out ...]
"""
import argparse
import sys
from datetime import date
from functools import partial
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: F401
from app.db.session import SessionLocal
from app.models.instrument import Instrument
from app.models.price import Price
from app.services.backtest_service import (
    BacktestConfig,
    _compute_metrics,
    compute_free_float_weights,
    compute_regime_exposure,
    run_momentum_backtest,
)
from app.services.factor_screen_service import ScreenConfig, screen_by_ebitda_peg

GUIDE_LINES = [
    "* 백테스팅 기간은 2020.1.1 ~ 2026.4.30 로 맞춰주시기 바랍니다.",
    '* 리밸런싱발생내역의 자산종류명은 알고리즘설명서의 "편입자산 종류 및 특징"의 자산종류명과 같아야 합니다.',
    "* 시계열 시트에 알고리즘의 시계열과 BM의 시계열을 함께 넣어주시기 바랍니다.",
    "* 날짜는 영업일 기준으로 작성해주세요.",
]

INDEX_LABEL_MAP = {"KOSPI": "코스피전체", "KOSPI200": "코스피200", "KOSDAQ": "코스닥전체", "KOSDAQ150": "코스닥150"}
BM_LABEL_MAP = {
    "KOSPI": "KOSPI(BM)",
    "KOSPI200": "KOSPI200(BM)",
    "KOSDAQ": "KOSDAQ(BM)",
    "KOSDAQ150": "KOSDAQ150TR(BM)",
}
GROUP_FIELD_LABEL = {"krx_sector": "업종", "sector": "섹터", "industry": "산업"}
REBALANCE_REASON = "정기리밸런싱"
PCT_FMT = "0.0%"


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--index", choices=list(INDEX_LABEL_MAP), default="KOSDAQ150")
    p.add_argument("--start", type=date.fromisoformat, default=date(2020, 1, 1))
    p.add_argument("--end", type=date.fromisoformat, default=date(2026, 7, 31))
    p.add_argument("--top-n", type=int, default=20)
    p.add_argument("--max-per-sector", type=int, default=None)
    p.add_argument("--screen-top-pct", type=float, default=0.5)
    p.add_argument("--min-sector-size", type=int, default=5)
    p.add_argument("--ttm-lag-days", type=int, default=90)
    p.add_argument("--consensus-lag-days", type=int, default=0)
    p.add_argument("--peg-min", type=float, default=0.0)
    p.add_argument("--stop-loss-pct", type=float, default=0.10)
    p.add_argument("--stop-loss-execution", choices=["close", "next_open"], default="next_open")
    p.add_argument("--ma-window-days", type=int, default=200)
    p.add_argument("--bear-exposure", type=float, default=0.5)
    p.add_argument("--weighting", choices=["equal", "free_float"], default="free_float")
    p.add_argument("--group-field", choices=["none", "krx_sector", "sector", "industry"], default="none")
    p.add_argument("--max-weight", type=float, default=0.30)
    p.add_argument("--max-group-weight", type=float, default=0.50)
    p.add_argument("--out", default=None)
    return p.parse_args()


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("작성가이드")
    for i, line in enumerate(GUIDE_LINES, start=3):
        ws.cell(row=i, column=2, value=line)


def build_timeseries_sheet(wb: Workbook, dates: list[date], mp_values: list[float], bm_values: list[float], bm_label: str):
    ws = wb.create_sheet("시계열")
    ws.cell(row=1, column=2, value="mp")
    ws.cell(row=1, column=3, value=bm_label)
    for i, (d, mp, bm) in enumerate(zip(dates, mp_values, bm_values), start=2):
        ws.cell(row=i, column=1, value=d).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=mp)
        ws.cell(row=i, column=3, value=bm)


def build_rebalance_sheet(wb: Workbook, rebalances: list[dict], instruments_by_ticker: dict[str, Instrument], asset_label: str):
    ws = wb.create_sheet("리밸런싱발생내역")

    tickers: list[str] = []
    seen = set()
    for r in rebalances:
        for t in r["holdings"]:
            if t not in seen:
                seen.add(t)
                tickers.append(t)

    n = len(tickers)
    reason_col = n + 2

    ws.cell(row=1, column=1, value="자산종류")
    ws.cell(row=1, column=2, value=asset_label)
    if n > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.cell(row=1, column=reason_col, value="리밸런싱 사유")
    ws.merge_cells(start_row=1, start_column=reason_col, end_row=3, end_column=reason_col)
    ws.cell(row=1, column=reason_col).alignment = Alignment(vertical="center")

    ws.cell(row=2, column=1, value="종목명")
    ws.cell(row=3, column=1, value="업종")
    for j, ticker in enumerate(tickers, start=2):
        inst = instruments_by_ticker[ticker]
        ws.cell(row=2, column=j, value=inst.name)
        ws.cell(row=3, column=j, value=inst.industry or inst.krx_sector)

    col_of_ticker = {t: j for j, t in enumerate(tickers, start=2)}
    for i, r in enumerate(rebalances, start=4):
        ws.cell(row=i, column=1, value=r["date"]).number_format = "yyyy-mm-dd"
        for t in r["holdings"]:
            cell = ws.cell(row=i, column=col_of_ticker[t], value=r["weights"][t])
            cell.number_format = PCT_FMT
        ws.cell(row=i, column=reason_col, value=REBALANCE_REASON)

    ws.column_dimensions[get_column_letter(1)].width = 12


def main():
    args = parse_args()
    benchmark_ticker = args.index
    benchmark_label = BM_LABEL_MAP[args.index]
    index_label = INDEX_LABEL_MAP[args.index]
    group_field = None if args.group_field == "none" else args.group_field

    config = BacktestConfig(
        index_name=args.index,
        top_n=args.top_n,
        max_per_sector=args.max_per_sector,
        sector_count_field=group_field,
        start_date=args.start,
        end_date=args.end,
    )
    screen_config = ScreenConfig(
        top_pct=args.screen_top_pct,
        min_sector_size=args.min_sector_size,
        ttm_lag_days=args.ttm_lag_days,
        consensus_lag_days=args.consensus_lag_days,
        peg_min=args.peg_min,
    )

    db = SessionLocal()
    screen_fn = partial(screen_by_ebitda_peg, config=screen_config, warn=print, info=lambda m: None)
    exposure_fn = partial(
        compute_regime_exposure,
        benchmark_ticker=benchmark_ticker,
        ma_window_days=args.ma_window_days,
        bull_exposure=1.0,
        bear_exposure=args.bear_exposure,
    )
    if args.weighting == "free_float":
        weight_fn = partial(
            compute_free_float_weights,
            max_weight=args.max_weight,
            group_field=group_field,
            max_group_weight=args.max_group_weight if group_field is not None else None,
        )
    else:
        weight_fn = None
    result = run_momentum_backtest(
        db,
        config,
        on_warning=print,
        on_info=lambda m: None,
        screen_fn=screen_fn,
        weight_fn=weight_fn,
        exposure_fn=exposure_fn,
        stop_loss_pct=args.stop_loss_pct,
        stop_loss_execution=args.stop_loss_execution,
    )

    if result.warnings:
        print(f"\n미분류 시계열단절 {len(result.warnings)}건이 남아있습니다. resolve_backtest_event.py로 먼저 해결하세요:")
        for w in result.warnings:
            print(f"  - {w}")
        db.close()
        sys.exit(1)

    official_dates = [d for d, _ in result.nav_series if d >= args.start]
    anchor_date = official_dates[0]
    nav_by_date = dict(result.nav_series)
    anchor_nav = nav_by_date[anchor_date]

    bench_rows = (
        db.query(Price.date, Price.close)
        .join(Instrument, Instrument.id == Price.instrument_id)
        .filter(Instrument.ticker == benchmark_ticker, Price.period == "D", Price.date.in_(official_dates))
        .all()
    )
    bench_by_date = {r.date: float(r.close) for r in bench_rows}
    missing_bench = [d for d in official_dates if d not in bench_by_date]
    if missing_bench:
        print(f"경고: 벤치마크에 없는 날짜 {len(missing_bench)}건 (예: {missing_bench[:5]})")
    anchor_bench = bench_by_date[anchor_date]

    mp_values = [nav_by_date[d] / anchor_nav * 100 for d in official_dates]
    bm_values = [bench_by_date[d] / anchor_bench * 100 for d in official_dates]

    all_tickers = {t for r in result.rebalances for t in r["holdings"]}
    instruments_by_ticker = {i.ticker: i for i in db.query(Instrument).filter(Instrument.ticker.in_(all_tickers)).all()}

    weighting_suffix = {
        "free_float": "(유동시총가중)",
        "equal": "(동일가중)",
    }[args.weighting]
    group_suffix = f"_{GROUP_FIELD_LABEL[group_field]}그룹{args.max_group_weight:.0%}캡" if group_field is not None else ""
    cap_suffix = f"_종목당최대{args.max_weight:.0%}" if args.max_weight is not None else ""
    sector_suffix = f"_섹터당{args.max_per_sector}개이하" if args.max_per_sector is not None else "_섹터캡없음"
    exec_label = "익일시가" if args.stop_loss_execution == "next_open" else "종가"
    asset_label = (
        f"{index_label} EBITDAPEG스크리닝모멘텀+월중손절{args.stop_loss_pct:.0%}({exec_label})+"
        f"레짐필터(약세시{args.bear_exposure:.0%}비중)" + sector_suffix + weighting_suffix + group_suffix + cap_suffix
    )

    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb)
    build_timeseries_sheet(wb, official_dates, mp_values, bm_values, bm_label=benchmark_label)
    build_rebalance_sheet(wb, result.rebalances, instruments_by_ticker, asset_label=asset_label)

    default_out = (
        f"../reference/(별첨 2) 백테스팅자료_{index_label} 실험10가중치비교_top{args.top_n}"
        f"{weighting_suffix}{group_suffix}.xlsx"
    )
    out_path = Path(args.out or default_out)
    wb.save(out_path)

    rebased = list(zip(official_dates, mp_values))
    m = _compute_metrics(rebased)
    db.close()

    print(f"\n저장 완료: {out_path}")
    print(f"  시계열: {len(official_dates)}행 ({official_dates[0]} ~ {official_dates[-1]})")
    print(f"  리밸런싱: {len(result.rebalances)}회, 누적종목수: {len(all_tickers)}개")
    print(
        f"  최종지수: {mp_values[-1]:.1f}  누적수익률: {m['cumulative_return']:+.1%}  CAGR: {m['cagr']:+.1%}  "
        f"연변동성: {m['annualized_volatility']:.1%}  MDD: {m['mdd']:.1%}  샤프: {m['sharpe']:.2f}"
    )


if __name__ == "__main__":
    main()
