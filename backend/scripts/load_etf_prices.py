"""DataGuide ETF 시계열 export를 instruments(asset_type='etf') + prices에 적재한다.

`load_wisefn_wide_prices.py`와 레이아웃 계열은 같지만 두 가지가 달라 별도로 둔다:

1. **헤더 행 위치가 export마다 다르다.** 요청 양식 재사용본은 Code가 8행인데 이 파일은
   7행이다. 행 번호를 고정하지 않고 A열에서 'Code'/'Name'/'Item Code'/'D A T E'를 찾는다.
2. **`asset_type='etf'`로 등록한다.** 주식과 섞이면 안 된다 — 우선주 판별(끝자리 규칙),
   KRX 업종분류 동기화, 지수 편입 로직이 전부 주식 전제다.

**배당조정(dividend_adjusted_prices)은 계산하지 않는다.** ETF 분배금 데이터가 없어서
`adj_close = close`로 채우면 분배금을 지급하는 ETF(국채·고배당 계열)의 총수익이 그만큼
과소평가된다. 잘못된 값을 넣느니 비워둔다 — 분배금이나 TR 시계열을 확보한 뒤에 채운다.
따라서 **ETF는 `prices`만 채워지고 `dividend_adjusted_prices`에는 안 들어간다.**
백테스트에서 ETF를 쓸 땐 이 점을 반드시 확인할 것.

수정주가(S100300)는 주식과 동일하게 `prices.close`(분할조정 계열)로 넣는다. ETF도
액면분할이 있다(예: 1/10 분할).

사용법:
  python scripts/load_etf_prices.py ~/Downloads/kretf.xlsx --dry-run
  python scripts/load_etf_prices.py ~/Downloads/kretf.xlsx
"""
import argparse
import datetime
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.market_holiday import MarketHoliday  # noqa: E402
from app.models.price import Price  # noqa: E402

# Item Code -> prices 컬럼. 지금은 수정주가만 받았지만 시가/고가/저가/거래량도 같은 표에서 온다.
ITEM_MAP = {
    "S100300": "close",   # 수정주가 (분할조정)
    "S100310": "open",
    "S100320": "high",
    "S100330": "low",
    "S100100": "raw_close",  # 종가(미조정) — prices에는 안 넣고 참고용
    "S101200": "volume",
}
PRICE_FIELDS = {"close", "open", "high", "low", "volume"}
BATCH_SIZE = 5000
ETF_MARKET = "유가증권시장"  # 국내 ETF는 전부 유가증권시장 상장


def find_header_rows(rows: list[tuple]) -> dict[str, int]:
    """A열 라벨로 헤더 행 위치를 찾는다 (export마다 1~2행씩 밀린다)."""
    want = {"code": "code", "name": "name", "item code": "item", "d a t e": "date"}
    found: dict[str, int] = {}
    for i, r in enumerate(rows[:30]):
        label = str(r[0]).strip().lower() if r and r[0] else ""
        if label in want:
            found[want[label]] = i
    missing = {"code", "item", "date"} - set(found)
    if missing:
        raise SystemExit(f"헤더 행을 찾지 못했습니다: {sorted(missing)} — 파일 형식을 확인하세요.")
    return found


def read_sheet(path: str, sheet: str | None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    h = find_header_rows(rows)
    codes, items = rows[h["code"]], rows[h["item"]]
    names = rows[h["name"]] if "name" in h else [None] * len(codes)

    # ticker -> {컬럼: 열번호}, ticker -> 이름
    columns: dict[str, dict[str, int]] = {}
    labels: dict[str, str] = {}
    skipped: list[str] = []
    for j in range(1, len(codes)):
        code, item = codes[j], items[j]
        if not code:
            continue
        ticker = str(code).strip().lstrip("A")
        field = ITEM_MAP.get(str(item).strip()) if item else None
        if field is None:
            skipped.append(f"{ticker}:{item}")
            continue
        if field in PRICE_FIELDS:
            columns.setdefault(ticker, {})[field] = j
        labels.setdefault(ticker, str(names[j]).strip() if names[j] else ticker)

    data: dict[str, dict[datetime.date, dict]] = {t: {} for t in columns}
    for row in rows[h["date"] + 1 :]:
        d = row[0]
        if d is None:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        elif not isinstance(d, datetime.date):
            continue
        for ticker, colmap in columns.items():
            vals = {c: float(row[j]) for c, j in colmap.items()
                    if row[j] is not None and isinstance(row[j], (int, float))}
            if vals.get("close"):  # 수정주가가 없는 날 = 상장 전이거나 거래 없음
                data[ticker][d] = vals
    return data, labels, skipped


def main(path: str, sheet: str | None, dry_run: bool):
    data, labels, skipped = read_sheet(path, sheet)
    if skipped:
        print(f"미매핑 Item Code {len(skipped)}열 무시: {skipped[:5]}")

    db = SessionLocal()
    existing = {t: (i, a) for t, i, a in db.query(Instrument.ticker, Instrument.id, Instrument.asset_type).all()}
    holidays = {r[0] for r in db.query(MarketHoliday.date).all()}

    print(f"\n{path}: ETF {len(data)}종목")
    conflict = [t for t in data if t in existing and existing[t][1] != "etf"]
    if conflict:
        print(f"  경고: 이미 다른 asset_type으로 등록된 티커 {len(conflict)}개 — 건너뜁니다: {conflict}")

    print(f"\n{'티커':8s} {'이름':32s} {'시작':12s} {'종료':12s} {'행수':>7s}  상태")
    to_load = {}
    for ticker, series in sorted(data.items(), key=lambda kv: min(kv[1]) if kv[1] else datetime.date.max):
        if not series or ticker in conflict:
            continue
        days = sorted(series)
        state = "기존" if ticker in existing else "신규등록"
        print(f"{ticker:8s} {labels.get(ticker, '')[:30]:32s} {str(days[0]):12s} {str(days[-1]):12s} "
              f"{len(days):>7,}  {state}")
        to_load[ticker] = series

    # 휴장일에 값이 있으면 벤더가 비거래일을 채운 것 — 주식 적재에서 겪은 유령행 문제와 같다.
    phantom = sum(1 for s in to_load.values() for d in s if d in holidays or d.weekday() >= 5)
    if phantom:
        print(f"\n  경고: 휴장일·주말에 값이 있는 행 {phantom:,}건 — 적재 시 제외합니다.")

    total = sum(len(s) for s in to_load.values())
    print(f"\n적재 대상: {len(to_load)}종목 / {total:,}행")
    if dry_run:
        print("--dry-run 이므로 변경하지 않고 종료합니다.")
        db.close()
        return

    rows: list[dict] = []
    created = 0
    for ticker, series in to_load.items():
        if ticker in existing:
            instrument_id = existing[ticker][0]
        else:
            inst = Instrument(ticker=ticker, name=labels.get(ticker, ticker),
                              asset_type="etf", market=ETF_MARKET)
            db.add(inst)
            db.flush()
            instrument_id = inst.id
            created += 1
        for d, vals in series.items():
            if d in holidays or d.weekday() >= 5:
                continue
            rows.append(dict(instrument_id=instrument_id, date=d, period="D",
                             volume=int(vals["volume"]) if "volume" in vals else None,
                             **{k: v for k, v in vals.items() if k in ("open", "high", "low", "close")}))

    for i in range(0, len(rows), BATCH_SIZE):
        stmt = pg_insert(Price).values(rows[i : i + BATCH_SIZE])
        stmt = stmt.on_conflict_do_update(
            index_elements=["instrument_id", "date", "period"],
            set_={c: getattr(stmt.excluded, c) for c in ("open", "high", "low", "close", "volume")},
        )
        db.execute(stmt)
    db.commit()

    print(f"\n신규 등록 {created}종목, prices {len(rows):,}행 적재")
    print("배당조정(dividend_adjusted_prices)은 계산하지 않았습니다 — 분배금 데이터가 없습니다.")
    db.close()


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("path")
    p.add_argument("--sheet", default=None, help="시트명 (기본: 첫 시트)")
    p.add_argument("--dry-run", action="store_true")
    a = p.parse_args()
    main(a.path, a.sheet, a.dry_run)
