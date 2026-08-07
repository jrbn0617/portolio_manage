"""유동주식시가총액 방식으로 재현한 KOSPI200/KOSDAQ150 지수를 일 단위로 계산해
엑셀로 저장한다. 기존에 적재된 실제 KOSPI200/KOSDAQ150 지수도 BM으로 함께 넣는다.

시가총액(비중) 갱신은 매월, 편입종목 변경은 반기 스냅샷(index_memberships)이
바뀔 때만 발생한다. 종목이 바뀌지 않은 달도 포함해 매 리밸런싱마다 편입종목·비중을
"KOSPI200_리밸런싱"/"KOSDAQ150_리밸런싱" 시트에 기록하고, "구분" 컬럼으로
종목변경/비중갱신을 구분한다.

사전 조건: 보유구간 도중 시계열단절(상장폐지/인수합병)이 미분류 상태면 경고 후 중단된다
— resolve_backtest_event.py로 먼저 해결해야 한다 (momentum 백테스트와 동일한
corporate_action_events를 공유하므로 이미 해결된 종목은 자동 반영된다).

--end을 생략하면 raw_close/월간펀더멘털/BM 지수가 공통으로 채워진 가장 최근 날짜까지
자동으로 계산한다(=실험용, "최근까지"). 그 달은 정식 월말 리밸런싱 없이 마지막 확정
비중으로 데이터가 있는 날짜까지만 시계열을 이어붙인다.

사용법: python scripts/export_free_float_index_excel.py [--start 2020-01-01] [--end 2026-04-30]
        [--out "../reference/유동주식시가총액지수_KOSPI200_KOSDAQ150.xlsx"]
"""
import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: F401
from app.db.session import SessionLocal
from app.models.index_membership import IndexMembership
from app.models.instrument import Instrument
from app.models.monthly_fundamental import MonthlyFundamental
from app.models.price import Price
from app.services.free_float_index_service import IndexBacktestConfig, IndexBacktestResult, run_free_float_index

YELLOW = "\033[93m"
CYAN = "\033[96m"
RESET = "\033[0m"

INDEX_NAMES = ["KOSPI200", "KOSDAQ150"]
BM_LABELS = {"KOSPI200": "KOSPI200(BM)", "KOSDAQ150": "KOSDAQ150TR(BM)"}
MCW_LABELS = {"KOSPI200": "KOSPI200(유동시총)", "KOSDAQ150": "KOSDAQ150(유동시총)"}
ASSET_TYPE_LABEL = {"KOSPI200": "코스피200 유동주식시가총액", "KOSDAQ150": "코스닥150 유동주식시가총액"}
PCT_FMT = "0.00%"


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--start", type=date.fromisoformat, default=date(2020, 1, 1))
    p.add_argument("--end", type=date.fromisoformat, default=None, help="생략 시 데이터가 공통으로 있는 최근 날짜")
    p.add_argument("--out", default="../reference/유동주식시가총액지수_KOSPI200_KOSDAQ150.xlsx")
    return p.parse_args()


def _latest_full_close_coverage_date(db, member_ids: list[int], upper_bound: date) -> date:
    """member_ids(현재 편입종목) 대부분(90%+)의 Price.close가 채워진, upper_bound 이하
    가장 최근 날짜. daily_update가 당일 일부 시장만 부분 적재된 날짜(예: 코스닥만 누락)를
    걸러내기 위함 — raw_close/월간펀더멘털은 채워졌어도 close가 비어있으면 일별 가치평가가
    깨진다."""
    rows = (
        db.query(Price.date, func.count(func.distinct(Price.instrument_id)))
        .filter(
            Price.instrument_id.in_(member_ids),
            Price.period == "D",
            Price.date <= upper_bound,
            Price.close.isnot(None),
        )
        .group_by(Price.date)
        .order_by(Price.date.desc())
        .limit(10)
        .all()
    )
    threshold = max(1, int(len(member_ids) * 0.9))
    for d, c in rows:
        if c >= threshold:
            return d
    return rows[-1][0] if rows else upper_bound


def detect_latest_available_date(db) -> date:
    """raw_close, 월간 펀더멘털(유동비율/상장주식수), BM 지수 종가, 그리고 각 지수
    현재 편입종목의 Price.close가 모두 채워진 가장 최근 날짜."""
    latest_price = db.query(func.max(Price.date)).filter(Price.period == "D", Price.raw_close.isnot(None)).scalar()
    latest_fund = db.query(func.max(MonthlyFundamental.date)).scalar()
    bm_maxes = [
        db.query(func.max(Price.date))
        .join(Instrument, Instrument.id == Price.instrument_id)
        .filter(Instrument.ticker == t, Price.period == "D")
        .scalar()
        for t in INDEX_NAMES
    ]
    upper_bound = min(latest_price, latest_fund, *bm_maxes)

    coverage_dates = []
    for index_name in INDEX_NAMES:
        snap = (
            db.query(IndexMembership.as_of_date)
            .filter(IndexMembership.index_name == index_name)
            .order_by(IndexMembership.as_of_date.desc())
            .first()[0]
        )
        members = [
            r[0]
            for r in db.query(IndexMembership.instrument_id)
            .filter(IndexMembership.index_name == index_name, IndexMembership.as_of_date == snap)
            .all()
        ]
        coverage_dates.append(_latest_full_close_coverage_date(db, members, upper_bound))

    return min(upper_bound, *coverage_dates)


def warn(msg: str) -> None:
    print(f"{YELLOW}{msg}{RESET}")


def info(msg: str) -> None:
    pass


def rebase_to_100(nav_series: list[tuple[date, float]], official_dates: list[date]) -> list[float]:
    nav_by_date = dict(nav_series)
    anchor = nav_by_date[official_dates[0]]
    return [nav_by_date[d] / anchor * 100 for d in official_dates]


def build_timeseries_sheet(wb: Workbook, dates: list[date], series: dict[str, list[float]]):
    ws = wb.create_sheet("시계열")
    headers = list(series.keys())
    for j, h in enumerate(headers, start=2):
        ws.cell(row=1, column=1, value="Date")
        ws.cell(row=1, column=j, value=h)
    for i, d in enumerate(dates, start=2):
        ws.cell(row=i, column=1, value=d).number_format = "yyyy-mm-dd"
        for j, h in enumerate(headers, start=2):
            ws.cell(row=i, column=j, value=series[h][i - 2])


def build_rebalance_sheet(
    wb: Workbook, sheet_name: str, asset_label: str, rebalances: list[dict], instruments_by_ticker: dict[str, Instrument]
):
    ws = wb.create_sheet(sheet_name)

    tickers: list[str] = []
    seen = set()
    for r in rebalances:
        for t in r["weights"]:
            if t not in seen:
                seen.add(t)
                tickers.append(t)

    n = len(tickers)
    reason_col = n + 2

    ws.cell(row=1, column=1, value="자산종류")
    ws.cell(row=1, column=2, value=asset_label)
    if n > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.cell(row=1, column=reason_col, value="구분")

    ws.cell(row=2, column=1, value="종목명")
    ws.cell(row=3, column=1, value="업종")
    for j, ticker in enumerate(tickers, start=2):
        inst = instruments_by_ticker[ticker]
        ws.cell(row=2, column=j, value=inst.name)
        ws.cell(row=3, column=j, value=inst.industry or inst.krx_sector)

    col_of_ticker = {t: j for j, t in enumerate(tickers, start=2)}
    for i, r in enumerate(rebalances, start=4):
        ws.cell(row=i, column=1, value=r["date"]).number_format = "yyyy-mm-dd"
        for t, w in r["weights"].items():
            cell = ws.cell(row=i, column=col_of_ticker[t], value=w)
            cell.number_format = PCT_FMT
        ws.cell(row=i, column=reason_col, value=r["reason"])

    ws.column_dimensions[get_column_letter(1)].width = 12


def main():
    args = parse_args()
    db = SessionLocal()

    end_date = args.end or detect_latest_available_date(db)
    if args.end is None:
        print(f"--end 미지정 → 데이터가 공통으로 있는 최근 날짜로 자동 설정: {end_date}")

    results: dict[str, IndexBacktestResult] = {}
    for index_name in INDEX_NAMES:
        print(f"{index_name} 유동주식시가총액 지수 계산 중...")
        cfg = IndexBacktestConfig(index_name=index_name, start_date=args.start, end_date=end_date)
        result = run_free_float_index(db, cfg, on_warning=warn, on_info=info)
        if result.warnings:
            print(
                f"\n{index_name}: 미분류 시계열단절 {len(result.warnings)}건이 남아있습니다. "
                "resolve_backtest_event.py로 먼저 해결하세요:"
            )
            for w in result.warnings:
                print(f"  - {w}")
            db.close()
            sys.exit(1)
        results[index_name] = result
        print(f"  완료 — 리밸런싱 {len(result.rebalances)}회, NAV {len(result.nav_series)}포인트")

    official_dates = [d for d, _ in results[INDEX_NAMES[0]].nav_series if d >= args.start]

    series: dict[str, list[float]] = {}
    for index_name in INDEX_NAMES:
        series[MCW_LABELS[index_name]] = rebase_to_100(results[index_name].nav_series, official_dates)

        bench_rows = (
            db.query(Price.date, Price.close)
            .join(Instrument, Instrument.id == Price.instrument_id)
            .filter(Instrument.ticker == index_name, Price.period == "D", Price.date.in_(official_dates))
            .all()
        )
        bench_by_date = {r.date: float(r.close) for r in bench_rows}
        missing_bench = [d for d in official_dates if d not in bench_by_date]
        if missing_bench:
            print(f"경고: {index_name} 벤치마크에 없는 날짜 {len(missing_bench)}건 (예: {missing_bench[:5]})")
        anchor_bench = bench_by_date[official_dates[0]]
        series[BM_LABELS[index_name]] = [bench_by_date.get(d, anchor_bench) / anchor_bench * 100 for d in official_dates]

    wb = Workbook()
    wb.remove(wb.active)
    build_timeseries_sheet(wb, official_dates, series)

    for index_name in INDEX_NAMES:
        all_tickers = {t for r in results[index_name].rebalances for t in r["weights"]}
        instruments_by_ticker = {i.ticker: i for i in db.query(Instrument).filter(Instrument.ticker.in_(all_tickers)).all()}
        build_rebalance_sheet(
            wb, f"{index_name}_리밸런싱", ASSET_TYPE_LABEL[index_name], results[index_name].rebalances, instruments_by_ticker
        )

    out_path = Path(args.out)
    wb.save(out_path)
    db.close()

    print(f"\n저장 완료: {out_path}")
    print(f"  시계열: {len(official_dates)}행 ({official_dates[0]} ~ {official_dates[-1]})")
    for index_name in INDEX_NAMES:
        r = results[index_name]
        changes = sum(1 for x in r.rebalances if x["universe_changed"])
        print(f"  {index_name}: 리밸런싱 {len(r.rebalances)}회 (종목변경 {changes}회)")

    for index_name in INDEX_NAMES:
        md = results[index_name].missing_data
        if not md:
            continue
        print(f"\n=== {index_name}: 데이터 부족으로 한 번이라도 제외된 종목 ({len(md)}개) ===")
        for row in md:
            print(
                f"  {row['ticker']} {row['name']}: {row['excluded_count']}회 제외 "
                f"({row['first_date']} ~ {row['last_date']}), 원인: {', '.join(row['reasons'])}"
            )


if __name__ == "__main__":
    main()
