"""2014~2018 DataGuide 백필 응답 4파일을 적재한다 (표본 확장용, 1회성).

`generate_history_backfill_requests.py`로 만든 요청에 대한 응답이다. 파일당 항목이 여럿이고
항목마다 500종목씩 7청크로 쪼개져 있다(시트명 = 항목명 + 청크번호).

  백필요청1_가격      D  수정주가·시가·고가·저가·종가·시가총액   -> prices, raw_closes
  백필요청2_월간팩터   M  상장주식수·유동주식비율·EBITDA 3종      -> monthly_fundamentals
  백필요청3_수급      D  기관합계/외국인/개인 x 수량·대금        -> investor_trading
  백필요청4_공매도    D  차입공매도 수량·금액 + 거래량·거래대금   -> short_selling

**시트명을 믿지 않고 10행(Item Code)으로 항목을 판별한다.** 요청 템플릿을 복사해 만든 응답은
시트명이 엉뚱하게 남아 있는 전례가 있다(load_short_selling_backfill.py 주의 1).

**단위 환산** — 기존 적재분과 스케일을 맞춘다:
  Shares thou -> x1,000 (순매수수량)   Local mn -> x1,000,000 (순매수대금)
  Local thou  -> x1,000 (EBITDA, 차입공매도금액)
  유동주식비율은 %(57·100) 스케일 그대로, 상장주식수·거래량·거래대금·시가총액은 그대로.

**유령행 차단**: WISEfn은 비거래일을 앞값으로 채워서 준다(과거에 prices에 177,706행이 들어와
문제가 됐다). `market_holidays`에는 2014~2017년이 비어 있어 그것만으로는 못 막는다. 대신
**KRX 원천인 ETF 시세(prices.asset_type='etf')에서 실제 거래일 달력을 뽑아** 필터로 쓴다 —
외부 API 호출이 필요 없다. 같은 달력으로 `market_holidays`의 2014~2017년 공백도 채운다
(백테스트 엔진의 `get_trading_days()`가 이 테이블을 쓴다).

**기존 데이터 보호**: 2018-12-28부터는 이미 적재돼 있다. 전 구간 ON CONFLICT DO NOTHING이라
기존 행은 건드리지 않는다.

메모리: 전체를 한 번에 올리면 2천만 행이라 못 버틴다. **청크(500종목)마다** 항목 시트를 묶어
long 포맷(ticker, date, item, value)으로 COPY한 뒤 Postgres에서 피벗해 본 테이블에 넣는다.

사용법:
  python scripts/load_history_backfill.py --dry-run
  python scripts/load_history_backfill.py
  python scripts/load_history_backfill.py --only 가격 수급
"""
import argparse
import datetime
import io
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
REFERENCE_DIR = BACKEND_DIR.parent / "reference"
sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import text  # noqa: E402

from app.db.base import Base  # noqa: E402,F401
from app.db.session import engine  # noqa: E402
from app.services.instrument_rules import is_preferred  # noqa: E402

CODE_ROW, ITEM_ROW, DATA_START_ROW = 8, 10, 15  # 1-indexed

# 파일 구분 -> (파일명 접두어, 적재 함수명)
GROUPS = ["가격", "월간팩터", "수급", "공매도"]

MONTHLY_METRIC = {
    "S101500": ("shares_outstanding_monthly", 1.0),
    "S102060": ("free_float_ratio", 1.0),          # % 스케일 그대로
    "M123005.M": ("ebitda_ttm", 1_000.0),          # Local thou -> 원
    "E123060.M": ("ebitda_fwd_12m", 1_000.0),
    "E331060.M": ("ev_ebitda_fwd_12m", 1.0),       # 배수
}
FLOW_ITEM = {
    "U110310": ("기관합계", "net_volume", 1_000.0),
    "U130310": ("외국인", "net_volume", 1_000.0),
    "U140310": ("개인", "net_volume", 1_000.0),
    "U110320": ("기관합계", "net_value", 1_000_000.0),
    "U130320": ("외국인", "net_value", 1_000_000.0),
    "U140320": ("개인", "net_value", 1_000_000.0),
}
SHORT_SCALE = {"S102310": 1.0, "S102340": 1_000.0, "S100900": 1.0, "S101200": 1.0}


def find_files() -> dict[str, Path]:
    out = {}
    for n in sorted(REFERENCE_DIR.iterdir()):
        nfc = unicodedata.normalize("NFC", n.name)
        if not (nfc.startswith("백필요청") and nfc.endswith(".xlsx")):
            continue
        for g in GROUPS:
            if g in nfc:
                out[g] = n
    return out


def chunk_of(sheet: str) -> str:
    m = re.search(r"(\d+)$", unicodedata.normalize("NFC", sheet))
    return m.group(1) if m else "0"


def stream_sheet(ws):
    """(item_code, [(ticker, date, value), ...]) — 값이 있는 셀만."""
    rows = ws.iter_rows(values_only=True)
    header = []
    for i, r in enumerate(rows, 1):
        header.append(r)
        if i >= DATA_START_ROW - 1:
            break
    codes, items = header[CODE_ROW - 1], header[ITEM_ROW - 1]
    item = next((str(v).strip() for v in items[1:] if v), None)
    tickers = {}
    for j in range(1, len(codes)):
        c = codes[j]
        if not c:
            continue
        t = str(c).strip().lstrip("A")
        if not is_preferred(t):
            tickers[j] = t
    for r in rows:
        d = r[0]
        if d is None:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        elif not isinstance(d, datetime.date):
            continue
        for j, t in tickers.items():
            if j < len(r) and r[j] is not None and isinstance(r[j], (int, float)):
                yield item, t, d, float(r[j])


def load_group(db, group: str, path: Path, dry_run: bool) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    by_chunk = defaultdict(list)
    for s in wb.sheetnames:
        by_chunk[chunk_of(s)].append(s)
    print(f"\n[{group}] {unicodedata.normalize('NFC', path.name)} — 시트 {len(wb.sheetnames)}, 청크 {len(by_chunk)}")

    stats = defaultdict(int)
    for ci, (chunk, sheets) in enumerate(sorted(by_chunk.items(), key=lambda kv: int(kv[0])), 1):
        db.execute(text("truncate stage"))
        buf = io.StringIO()
        n = 0
        for s in sheets:
            for item, t, d, v in stream_sheet(wb[s]):
                buf.write(f"{t}\t{d}\t{item}\t{v}\n")
                n += 1
        buf.seek(0)
        with db.connection.cursor() as cur:
            cur.copy_expert("copy stage (ticker, d, item, val) from stdin with (format text)", buf)
        del buf
        stats["셀"] += n

        if dry_run:
            print(f"  청크 {chunk}: {n:,}셀 (dry-run, 적재 안 함)")
            continue

        stats["행"] += INSERTS[group](db)
        db.commit()
        print(f"  청크 {chunk}: {n:,}셀 -> 누적 {stats['행']:,}행")
    wb.close()
    return stats


def _ins_price(db) -> int:
    r = db.execute(text("""
        insert into prices (instrument_id, date, period, open, high, low, close, raw_close, market_cap)
        select i.id, s.d, 'D',
               max(val) filter (where item='S100310'),
               max(val) filter (where item='S100320'),
               max(val) filter (where item='S100330'),
               max(val) filter (where item='S100300'),
               max(val) filter (where item='S100100'),
               (max(val) filter (where item='S102100'))::bigint
        from stage s
        join instruments i on i.ticker = s.ticker and i.asset_type = 'stock'
        join trading_day t on t.d = s.d
        group by i.id, s.d
        having max(val) filter (where item='S100300') > 0
        on conflict (instrument_id, date, period) do nothing"""))
    db.execute(text("""
        insert into raw_closes (instrument_id, date, close)
        select i.id, s.d, max(val) filter (where item='S100100')
        from stage s
        join instruments i on i.ticker = s.ticker and i.asset_type = 'stock'
        join trading_day t on t.d = s.d
        group by i.id, s.d
        having max(val) filter (where item='S100100') > 0
        on conflict (instrument_id, date) do nothing"""))
    return r.rowcount


def _ins_monthly(db) -> int:
    total = 0
    for item, (metric, scale) in MONTHLY_METRIC.items():
        r = db.execute(text("""
            insert into monthly_fundamentals (instrument_id, date, metric, value)
            select i.id, s.d, :m, s.val * :k
            from stage s join instruments i on i.ticker = s.ticker and i.asset_type = 'stock'
            where s.item = :it
            on conflict (instrument_id, date, metric) do nothing"""),
            {"m": metric, "k": scale, "it": item})
        total += r.rowcount
    return total


def _ins_flow(db) -> int:
    for inv in ("기관합계", "외국인", "개인"):
        vol = [k for k, v in FLOW_ITEM.items() if v[0] == inv and v[1] == "net_volume"][0]
        val = [k for k, v in FLOW_ITEM.items() if v[0] == inv and v[1] == "net_value"][0]
        db.execute(text(f"""
            insert into investor_trading (instrument_id, date, investor_type, net_volume, net_value)
            select i.id, s.d, :inv,
                   (max(val) filter (where item='{vol}') * 1000)::bigint,
                   (max(val) filter (where item='{val}') * 1000000)::bigint
            from stage s
            join instruments i on i.ticker = s.ticker and i.asset_type = 'stock'
            join trading_day t on t.d = s.d
            where s.item in ('{vol}','{val}')
            group by i.id, s.d
            on conflict (instrument_id, date, investor_type) do nothing"""), {"inv": inv})
    return db.execute(text("select count(*) from stage")).scalar() // 2


def _ins_short(db) -> int:
    r = db.execute(text("""
        insert into short_selling (instrument_id, date, short_volume, short_value,
                                   total_volume, total_value, volume_ratio, value_ratio)
        select i.id, s.d,
               (max(val) filter (where item='S102310'))::bigint,
               (max(val) filter (where item='S102340') * 1000)::bigint,
               (max(val) filter (where item='S100900'))::bigint,
               (max(val) filter (where item='S101200'))::bigint,
               case when max(val) filter (where item='S100900') > 0
                    then round((max(val) filter (where item='S102310'))::numeric
                             / (max(val) filter (where item='S100900'))::numeric * 100, 4) end,
               case when max(val) filter (where item='S101200') > 0
                    then round((max(val) filter (where item='S102340') * 1000)::numeric
                             / (max(val) filter (where item='S101200'))::numeric * 100, 4) end
        from stage s
        join instruments i on i.ticker = s.ticker and i.asset_type = 'stock'
        join trading_day t on t.d = s.d
        group by i.id, s.d
        on conflict (instrument_id, date) do nothing"""))
    return r.rowcount


INSERTS = {"가격": _ins_price, "월간팩터": _ins_monthly, "수급": _ins_flow, "공매도": _ins_short}


def main(only: list[str] | None, dry_run: bool):
    files = find_files()
    todo = [g for g in GROUPS if g in files and (not only or g in only)]
    if not todo:
        raise SystemExit(f"대상 파일이 없습니다. reference/에서 찾은 것: {list(files)}")

    # Session 대신 Connection 을 쓴다 — Session.commit() 은 커넥션을 풀에 반납할 수 있고
    # 그러면 임시테이블(stage/trading_day)이 통째로 사라진다. Connection 은 commit 후에도
    # 같은 커넥션에서 새 트랜잭션을 시작하므로 임시테이블이 유지된다.
    db = engine.connect()
    db.execute(text("create temp table stage (ticker varchar(32), d date, item varchar(16), val numeric)"))
    db.execute(text("create index on stage (item)"))

    # KRX 원천(ETF 시세)에서 실제 거래일 달력을 뽑는다 — 벤더 유령행 차단용
    db.execute(text("""
        create temp table trading_day as
        select distinct p.date d from prices p join instruments i on i.id = p.instrument_id
        where i.asset_type = 'etf' and p.period = 'D'"""))
    db.execute(text("create unique index on trading_day (d)"))
    lo, hi, n = db.execute(text("select min(d), max(d), count(*) from trading_day")).fetchone()
    print(f"거래일 달력: {n:,}일 ({lo} ~ {hi}) — KRX ETF 시세에서 도출")

    for g in todo:
        load_group(db, g, files[g], dry_run)

    if not dry_run:
        # prices.volume 은 가격 파일에 없다 — 공매도 파일의 거래량으로 채운다
        if "공매도" in todo:
            r = db.execute(text("""
                update prices p set volume = s.total_volume
                from short_selling s
                where s.instrument_id = p.instrument_id and s.date = p.date
                  and p.period = 'D' and p.volume is null and s.total_volume is not null"""))
            db.commit()
            print(f"\nprices.volume 보정 {r.rowcount:,}행 (공매도 파일의 거래량)")

        # market_holidays 2014~2017 공백 채우기 — 백테스트 엔진의 get_trading_days()가 쓴다
        # 주의: generate_series 별칭을 trading_day 의 컬럼명(d)과 겹치게 두면 서브쿼리 안에서
        # 바깥 d 가 안쪽 t.d 로 해석돼 `t.d = t.d`(항상 참)가 되고 NOT EXISTS 가 전부 거짓이 된다.
        # 실제로 그 버그로 78건이 0건으로 나갔다. 별칭을 cal(day) 로 분리한다.
        r = db.execute(text("""
            insert into market_holidays (date)
            select cal.day::date from generate_series(
                     (select min(d) from trading_day where d >= '2014-01-01'),
                     '2018-12-31'::date, '1 day') cal(day)
            where extract(dow from cal.day) between 1 and 5
              and not exists (select 1 from trading_day t where t.d = cal.day::date)
              and not exists (select 1 from market_holidays m where m.date = cal.day::date)"""))
        db.commit()
        print(f"market_holidays 신규 {r.rowcount:,}건 (2014~2018 공백)")

    report(db)
    db.close()


def report(db):
    print("\n" + "=" * 74)
    print("적재 후 현황 (2019-01-01 이전 = 이번 백필 구간)")
    print("=" * 74)
    rows = db.execute(text("""
        select 'prices(D)', count(*) filter (where p.date < '2019-01-01'), min(p.date), max(p.date)
          from prices p join instruments i on i.id=p.instrument_id
          where i.asset_type='stock' and p.period='D'
        union all select 'raw_closes', count(*) filter (where r.date < '2019-01-01'), min(r.date), max(r.date)
          from raw_closes r join instruments i on i.id=r.instrument_id where i.asset_type='stock'
        union all select 'monthly_fundamentals', count(*) filter (where m.date < '2019-01-01'), min(m.date), max(m.date)
          from monthly_fundamentals m join instruments i on i.id=m.instrument_id where i.asset_type='stock'
        union all select 'investor_trading', count(*) filter (where t.date < '2019-01-01'), min(t.date), max(t.date)
          from investor_trading t join instruments i on i.id=t.instrument_id where i.asset_type='stock'
        union all select 'short_selling', count(*) filter (where s.date < '2019-01-01'), min(s.date), max(s.date)
          from short_selling s join instruments i on i.id=s.instrument_id where i.asset_type='stock'
    """)).fetchall()
    for name, n, lo, hi in rows:
        print(f"  {name:22s} 백필분 {n:>10,}행   전체범위 {lo} ~ {hi}")
    print("\n  배당조정(dividend_adjusted_prices)은 아직 재계산하지 않았습니다 —")
    print("  backfill_dividend_adjusted_prices.py 를 별도로 돌려야 합니다.")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--only", nargs="*", choices=GROUPS)
    p.add_argument("--dry-run", action="store_true")
    a = p.parse_args()
    main(a.only, a.dry_run)
