"""DataGuide 백필 응답의 시가총액(S102100)을 prices.market_cap에 채운다 (1회성).

prices.market_cap은 daily_update.py(L150)가 pykrx 시가총액으로 이미 적재 중이라
2026-08-04 이후는 자동으로 들어오지만, 그 이전 과거 구간이 비어 있었다. 이 스크립트가
그 구간을 메운다.

INSERT가 아니라 UPDATE만 한다 — prices.close가 NOT NULL이라 시가총액만 있는 행을 새로
만들 수는 없고, 이미 있는 일봉 행에 값을 덧붙이는 성격이기 때문이다. 또 이미 값이 있는
행(= pykrx 적재분)은 건드리지 않는다.

단위: 원본이 "Local"(원)이라 변환 없이 그대로 저장한다. 실측 확인 —
삼성전자 2026-07-31 = 1,534,648,134,600,000원으로 pykrx 값과 정확히 일치.

시트명은 요청 템플릿 잔재라 무의미하므로 행10(Item Code)으로 항목을 판별한다.

사용법:
  python scripts/load_market_cap_backfill.py [../reference/backfill_data_response_3.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
from psycopg2.extras import execute_values
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402

ITEM_CODE = "S102100"
CODE_ROW = 8
ITEM_CODE_ROW = 10
DATA_START_ROW = 15
INSERT_PAGE = 10_000


def _clean_ticker(code) -> str:
    code = str(code).strip()
    return code[1:] if code.startswith("A") else code


def main(path: str):
    print(f"reading {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    db = SessionLocal()
    instruments_by_ticker = {t: i for t, i in db.query(Instrument.ticker, Instrument.id).all()}

    conn = db.connection().connection  # psycopg2 raw connection
    cur = conn.cursor()
    cur.execute("""
        create temporary table tmp_mcap (
            instrument_id integer not null,
            date date not null,
            market_cap bigint not null
        ) on commit drop""")

    seen: set[tuple[int, object]] = set()
    total = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(min_row=CODE_ROW, max_row=ITEM_CODE_ROW, values_only=True)
        row_code = next(it)
        next(it)
        if next(it)[1] != ITEM_CODE:
            continue
        tickers = [_clean_ticker(c) for c in row_code[1:] if c]

        buf = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            d = row[0]
            if d is None:
                continue
            d = d.date() if hasattr(d, "date") else d
            for ticker, cell in zip(tickers, row[1:]):
                if cell is None or isinstance(cell, str):
                    continue
                iid = instruments_by_ticker.get(ticker)
                if iid is None:
                    continue
                key = (iid, d)
                if key in seen:      # Sheet1에 중복 종목이 섞여 있어 (iid,date) 단위로 1건만
                    continue
                seen.add(key)
                buf.append((iid, d, round(float(cell))))

        for i in range(0, len(buf), INSERT_PAGE):
            execute_values(cur, "insert into tmp_mcap (instrument_id, date, market_cap) values %s",
                           buf[i : i + INSERT_PAGE], page_size=INSERT_PAGE)
        total += len(buf)
        print(f"  {sn}: {len(buf):,}행 적재 (누적 {total:,})")

    wb.close()

    print("prices.market_cap 갱신 중 (기존 값 보존, NULL만 채움) ...")
    cur.execute("""
        update prices p set market_cap = t.market_cap
          from tmp_mcap t
         where p.instrument_id = t.instrument_id
           and p.date = t.date
           and p.period = 'D'
           and p.market_cap is null""")
    updated = cur.rowcount
    conn.commit()
    print(f"  {updated:,}행 갱신")

    r = db.execute(text("""
        select count(*), count(market_cap) from prices where period='D'""")).fetchone()
    print(f"\n최종: prices(D) {r[0]:,}행 중 market_cap {r[1]:,}행 ({r[1]/r[0]*100:.1f}%)")
    db.close()
    print("완료.")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "../reference/backfill_data_response_3.xlsx"
    main(target)
