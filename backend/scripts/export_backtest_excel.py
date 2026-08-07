"""백테스트 결과를 reference/(별첨 2) 백테스팅자료_회사명(예시).xlsx 포맷으로 출력한다.

시트 3개: 작성가이드(원본 문구 그대로), 시계열(전략 NAV vs 벤치마크, 공식
기간 첫 거래일 기준 100 리베이스), 리밸런싱발생내역(리밸런싱별 종목 비중 매트릭스).

사전 조건: run_backtest.py로 같은 기간을 먼저 실행해 미분류 시계열단절
경고가 0건인지 확인해야 한다 — 있으면 resolve_backtest_event.py로 먼저 해결.

사용법: python scripts/export_backtest_excel.py [--start 2020-01-01] [--end 2026-04-30]
        [--out "../reference/(별첨 2) 백테스팅자료_코스닥150모멘텀.xlsx"]
"""
import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: F401
from app.db.session import SessionLocal
from app.models.instrument import Instrument
from app.models.price import Price
from app.services.backtest_service import BacktestConfig, run_momentum_backtest

GUIDE_LINES = [
    "* 백테스팅 기간은 2020.1.1 ~ 2026.4.30 로 맞춰주시기 바랍니다.",
    '* 리밸런싱발생내역의 자산종류명은 알고리즘설명서의 "편입자산 종류 및 특징"의 자산종류명과 같아야 합니다.',
    "* 시계열 시트에 알고리즘의 시계열과 BM의 시계열을 함께 넣어주시기 바랍니다.",
    "* 날짜는 영업일 기준으로 작성해주세요.",
]

ASSET_TYPE_LABEL = "코스닥150 모멘텀"
BENCHMARK_TICKER = "KOSDAQ150"
BENCHMARK_LABEL = "KOSDAQ150TR(BM)"
REBALANCE_REASON = "정기리밸런싱"
PCT_FMT = "0.0%"


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--start", type=date.fromisoformat, default=date(2020, 1, 1))
    p.add_argument("--end", type=date.fromisoformat, default=date(2026, 4, 30))
    p.add_argument("--top-n", type=int, default=10)
    p.add_argument("--out", default="../reference/(별첨 2) 백테스팅자료_코스닥150모멘텀.xlsx")
    return p.parse_args()


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("작성가이드")
    for i, line in enumerate(GUIDE_LINES, start=3):
        ws.cell(row=i, column=2, value=line)


def build_timeseries_sheet(wb: Workbook, dates: list[date], mp_values: list[float], bm_values: list[float]):
    ws = wb.create_sheet("시계열")
    ws.cell(row=1, column=2, value="mp")
    ws.cell(row=1, column=3, value=BENCHMARK_LABEL)
    for i, (d, mp, bm) in enumerate(zip(dates, mp_values, bm_values), start=2):
        ws.cell(row=i, column=1, value=d).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=mp)
        ws.cell(row=i, column=3, value=bm)


def build_rebalance_sheet(wb: Workbook, rebalances: list[dict], instruments_by_ticker: dict[str, Instrument]):
    ws = wb.create_sheet("리밸런싱발생내역")

    tickers: list[str] = []
    seen = set()
    for r in rebalances:
        for t in r["holdings"]:
            if t not in seen:
                seen.add(t)
                tickers.append(t)

    n = len(tickers)
    reason_col = n + 2  # A=1(라벨), B..=종목, 마지막=리밸런싱 사유

    ws.cell(row=1, column=1, value="자산종류")
    ws.cell(row=1, column=2, value=ASSET_TYPE_LABEL)
    if n > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.cell(row=1, column=reason_col, value="리밸런싱 사유")
    ws.merge_cells(start_row=1, start_column=reason_col, end_row=3, end_column=reason_col)
    ws.cell(row=1, column=reason_col).alignment = Alignment(vertical="center")

    ws.cell(row=2, column=1, value="종목명")
    ws.cell(row=3, column=1, value="업종")
    for j, ticker in enumerate(tickers, start=2):
        inst = instruments_by_ticker[ticker]
        ws.cell(row=2, column=j, value=inst.name)
        ws.cell(row=3, column=j, value=inst.industry or inst.krx_sector)

    col_of_ticker = {t: j for j, t in enumerate(tickers, start=2)}
    for i, r in enumerate(rebalances, start=4):
        ws.cell(row=i, column=1, value=r["date"]).number_format = "yyyy-mm-dd"
        for t in r["holdings"]:
            cell = ws.cell(row=i, column=col_of_ticker[t], value=r["weights"][t])
            cell.number_format = PCT_FMT
        ws.cell(row=i, column=reason_col, value=REBALANCE_REASON)

    ws.column_dimensions[get_column_letter(1)].width = 12


def main():
    args = parse_args()
    config = BacktestConfig(
        index_name="KOSDAQ150", top_n=args.top_n, start_date=args.start, end_date=args.end
    )

    db = SessionLocal()
    result = run_momentum_backtest(db, config, on_warning=print, on_info=lambda m: None)

    if result.warnings:
        print(f"\n미분류 시계열단절 {len(result.warnings)}건이 남아있습니다. resolve_backtest_event.py로 먼저 해결하세요:")
        for w in result.warnings:
            print(f"  - {w}")
        db.close()
        sys.exit(1)

    # 공식 기간 첫 거래일(예: 2020-01-02) 기준으로 100 리베이스, 그 이전(초기 형성일)은 시계열 시트에서 제외
    official_dates = [d for d, _ in result.nav_series if d >= args.start]
    anchor_date = official_dates[0]
    nav_by_date = dict(result.nav_series)
    anchor_nav = nav_by_date[anchor_date]

    bench_rows = (
        db.query(Price.date, Price.close)
        .join(Instrument, Instrument.id == Price.instrument_id)
        .filter(Instrument.ticker == BENCHMARK_TICKER, Price.period == "D", Price.date.in_(official_dates))
        .all()
    )
    bench_by_date = {r.date: float(r.close) for r in bench_rows}
    missing_bench = [d for d in official_dates if d not in bench_by_date]
    if missing_bench:
        print(f"경고: 벤치마크에 없는 날짜 {len(missing_bench)}건 (예: {missing_bench[:5]})")
    anchor_bench = bench_by_date[anchor_date]

    mp_values = [nav_by_date[d] / anchor_nav * 100 for d in official_dates]
    bm_values = [bench_by_date[d] / anchor_bench * 100 for d in official_dates]

    all_tickers = {t for r in result.rebalances for t in r["holdings"]}
    instruments_by_ticker = {
        i.ticker: i for i in db.query(Instrument).filter(Instrument.ticker.in_(all_tickers)).all()
    }

    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb)
    build_timeseries_sheet(wb, official_dates, mp_values, bm_values)
    build_rebalance_sheet(wb, result.rebalances, instruments_by_ticker)

    out_path = Path(args.out)
    wb.save(out_path)
    db.close()

    print(f"\n저장 완료: {out_path}")
    print(f"  시계열: {len(official_dates)}행 ({official_dates[0]} ~ {official_dates[-1]})")
    print(f"  리밸런싱: {len(result.rebalances)}회, 누적종목수: {len(all_tickers)}개")


if __name__ == "__main__":
    main()
