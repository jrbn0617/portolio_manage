"""monthly_fundamentals의 **결측 구간만** 다시 받기 위한 DataGuide 요청 양식을 만든다.

`generate_history_backfill_requests.py`가 "구간 전체를 처음 받는" 용도라면, 이쪽은
이미 적재된 데이터를 훑다가 발견한 **구멍을 메우는** 용도다. 항목 하나당 파일 하나로
만든다 — `load_monthly_fundamentals.py`가 파일 단위로 metric을 받기 때문이다
(`load_monthly_fundamentals.py <파일> <metric>`).

현재 등록된 프리셋 (2026-08-13 실측)

  shares2019   상장주식수 2019-01 ~ 2019-11
      월 138~291종목뿐이다(정상 달 약 2,297종목 = 90% 결측). 2018-12는 2,134,
      2019-12는 2,309로 앞뒤는 멀쩡해 이 구간만 뚫려 있다.
      `compute_free_float_weights`가 `_latest_monthly_value`로 "as_of 이하 최근값"을
      쓰기 때문에 **에러 없이 조용히 과거 값으로 폴백한다.** 2020년 이후 백테스트는
      2019-12가 온전해 영향이 제한적이지만 홀드아웃(2015~2019)에는 정면으로 걸린다.

  evebitda2014  EV EBITDA(Fwd.12M) 2014-01 ~ 2018-12
      2019-01을 경계로 커버리지가 2배가 된다(2018-12 207종목 -> 2019-01 411종목).
      같은 백필 파일로 받은 EBITDA(Fwd.12M)는 경계에서 안 끊기는데 이 항목만 끊긴다.
      PEG 스크리닝은 TTM·Fwd·EV/EBITDA 셋이 다 있어야 종목을 통과시키므로
      **2015~2018 후보 풀이 절반으로 줄어든다.** 원인이 (a) 그 시절 컨센서스 자체가
      없어서인지 (b) 백필 요청 때 이 항목만 조건이 안 맞았는지 구분이 안 돼서 다시 받는다.
      같은 결과가 오면 (a)로 확정하고 홀드아웃 설계에 반영한다.

**요청 구간은 결측 구간보다 넓게 잡는다.** 이미 가진 달을 일부러 겹쳐 넣어, 받아온 값이
기존 DB와 일치하는지 먼저 대조한 뒤(단위·정의 확인) 적재한다.

**우선주·ETF는 넣지 않는다.** `generate_history_backfill_requests.fetch_universe('all')`은
`asset_type <> 'index'`만 걸러서 2026-08-12에 적재한 ETF 1,195개가 섞여 들어온다
(ETF는 상장주식수가 아니라 상장좌수라 항목이 맞지 않는다). 여기서는 `asset_type='stock'`
+ 종목코드 끝자리 규칙(`is_common_stock`)으로 제한한다.

사용법:
  python scripts/generate_monthly_gap_requests.py                    # 전체 프리셋
  python scripts/generate_monthly_gap_requests.py shares2019
  python scripts/generate_monthly_gap_requests.py --universe index
"""
import argparse
import datetime
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import text  # noqa: E402

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.services.instrument_rules import is_common_stock  # noqa: E402
from scripts.generate_history_backfill_requests import (  # noqa: E402
    CHUNK_SIZE,
    OUTPUT_DIR,
    chunked,
    fill_sheet,
    resolve_template,
)


@dataclass(frozen=True)
class Preset:
    key: str
    metric: str          # monthly_fundamentals.metric — 적재 시 그대로 쓴다
    sheet_prefix: str
    item_code: str
    unit: str
    base_date: str
    period_from: datetime.date   # 요청 구간(겹침 포함)
    period_to: datetime.date
    gap_from: datetime.date      # 실제 결측 구간(보고용)
    gap_to: datetime.date
    out_name: str


PRESETS = {
    p.key: p
    for p in [
        Preset("shares2019", "shares_outstanding_monthly", "상장주식수",
               "S101500", "Shares", "",
               datetime.date(2018, 12, 1), datetime.date(2019, 12, 31),
               datetime.date(2019, 1, 1), datetime.date(2019, 11, 30),
               "요청_상장주식수_2019.xlsx"),
        Preset("evebitda2014", "ev_ebitda_fwd_12m", "EVEBITDA",
               "E331060.M", "X", "",
               datetime.date(2014, 1, 1), datetime.date(2019, 3, 31),
               datetime.date(2013, 12, 1), datetime.date(2018, 12, 31),
               "요청_EVEBITDA_2014.xlsx"),
    ]
}

UNIVERSE_SQL = {
    "all": """
        select distinct i.ticker, i.name from instruments i
        where i.asset_type = 'stock'
          and (exists (select 1 from prices p where p.instrument_id=i.id and p.period='D')
               or exists (select 1 from index_memberships m where m.instrument_id=i.id))
        order by 1
    """,
    "index": """
        select distinct i.ticker, i.name from instruments i
        join index_memberships m on m.instrument_id = i.id
        where m.index_name in ('KOSPI200','KOSDAQ150') and i.asset_type = 'stock'
        order by 1
    """,
}


def fetch_stocks(db, scope: str) -> list[tuple[str, str]]:
    rows = db.execute(text(UNIVERSE_SQL[scope])).fetchall()
    return [(r.ticker, r.name or r.ticker) for r in rows if is_common_stock(r.ticker)]


def report_coverage(db, p: Preset) -> None:
    rows = db.execute(text("""
        select to_char(date,'YYYY-MM') m, count(*) n from monthly_fundamentals
        where metric = :metric and date between :a and :b group by 1 order by 1"""),
        {"metric": p.metric,
         "a": p.period_from - datetime.timedelta(days=62),
         "b": p.period_to + datetime.timedelta(days=31)}).fetchall()
    # 월말 날짜가 종목마다 하루이틀 갈리므로 **연월로 묶어야** 한다. date로 묶으면
    # 같은 달이 여러 그룹으로 쪼개져 중앙값이 터무니없이 작게 나온다.
    med = db.execute(text("""
        select percentile_disc(0.5) within group (order by n) from (
          select count(*) n from monthly_fundamentals where metric=:metric
          group by to_char(date,'YYYY-MM')) t"""),
        {"metric": p.metric}).scalar()
    print(f"  현재 커버리지 (정상 수준 약 {med:,}종목/월)")
    shown = [r for r in rows if r[0] >= (p.gap_from - datetime.timedelta(days=62)).strftime("%Y-%m")]
    for m, n in shown[-20:]:
        flag = "  <-- 결측" if p.gap_from.strftime("%Y-%m") <= m <= p.gap_to.strftime("%Y-%m") else ""
        print(f"    {m}  {n:>6,}{flag}")


def build(db, p: Preset, universe: list[tuple[str, str]]) -> Path:
    chunks = chunked(universe, CHUNK_SIZE)
    tpl_path, tpl_sheet = resolve_template()
    wb = openpyxl.load_workbook(tpl_path)
    sample = wb[tpl_sheet]

    keep = set()
    for i, chunk in enumerate(chunks, start=1):
        ws = wb.copy_worksheet(sample)
        ws.title = f"{p.sheet_prefix}{i}" if len(chunks) > 1 else p.sheet_prefix
        fill_sheet(ws, chunk, "M", p.period_from, p.period_to, p.item_code, p.unit, p.base_date)
        keep.add(ws.title)
    for name in [n for n in wb.sheetnames if n not in keep]:
        del wb[name]

    out = OUTPUT_DIR / p.out_name
    wb.save(out)
    return out


def main(keys: list[str], scope: str) -> None:
    db = SessionLocal()
    universe = fetch_stocks(db, scope)
    made = []
    for key in keys:
        p = PRESETS[key]
        print(f"\n[{p.key}] {p.sheet_prefix} · metric={p.metric}")
        report_coverage(db, p)
        out = build(db, p, universe)
        n_sheets = len(openpyxl.load_workbook(out).sheetnames)
        print(f"  생성: {out}")
        print(f"    유니버스 {len(universe):,}종목({scope}) · 시트 {n_sheets}장 "
              f"· Item {p.item_code} · Unit {p.unit} · Frequency M")
        print(f"    Period {p.period_from} ~ {p.period_to}  "
              f"(결측 {p.gap_from}~{p.gap_to} + 대조용 겹침)")
        made.append((out, p.metric))
    db.close()

    print("\n" + "=" * 78)
    print("받아온 뒤 (겹친 달이 기존 DB 값과 맞는지 먼저 대조하고 적재)")
    for out, metric in made:
        print(f'  python scripts/load_monthly_fundamentals.py "{out}" {metric}')


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    # nargs="*" 에 choices를 걸면 인자를 아예 안 줬을 때 argparse가 빈 리스트를 choices와
    # 대조해서 실패한다. choices 없이 받고 아래에서 직접 검증한다.
    ap.add_argument("presets", nargs="*", help=f"생략하면 전체. 가능: {', '.join(PRESETS)}")
    ap.add_argument("--universe", choices=["all", "index"], default="all",
                    help="all=DB 전체 보통주(코스피전체/코스닥전체 유니버스 대응)")
    a = ap.parse_args()
    keys = a.presets or list(PRESETS)
    unknown = [k for k in keys if k not in PRESETS]
    if unknown:
        ap.error(f"모르는 프리셋 {unknown} — 가능: {', '.join(PRESETS)}")
    main(keys, a.universe)
