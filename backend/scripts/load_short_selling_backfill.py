"""1회성 백필 응답(reference/backfill_data_response_1.xlsx)을 short_selling 테이블에 적재한다.
공매도 4종(차입공매도수량/금액, 전체 거래량/거래대금, 2018-12-28~)을 종목x일자로 병합해
한 행으로 만들고, 비중 2종은 계산해서 채운다. 정기 실행용이 아님.

주의 1 — 시트명을 믿으면 안 된다: 이 응답 파일은 요청 템플릿을 복사해 만든 탓에 시트명이
"U110320_*"(수급 백필 때 쓴 이름) 그대로 남아있고 실제 Item Code와 무관하다. 그래서
시트명 접두어가 아니라 **행10(Item Code)**을 읽어서 항목을 판별한다.

주의 2 — S101200(거래대금)은 청크2가 누락됐다: 청크1 시트가 두 번 들어있고 청크2(500종목,
A013310~)가 빠져 있어 거래대금만 2,481종목이다(나머지 3종은 2,981종목). 누락분의
total_value/value_ratio는 NULL로 남는다 — volume_ratio는 전종목 계산되므로 분석에는
지장이 없다. 필요하면 청크2만 재요청해서 이 스크립트를 다시 돌리면 채워진다.

주의 3 — 기존 pykrx 데이터는 보존한다: daily_update.py가 2026-08-04부터 pykrx로
short_selling을 적재 중이라 구간이 겹친다. load_investor_trading_backfill.py와 동일하게
ON CONFLICT DO NOTHING으로 기존 행은 건드리지 않는다(정정분은 일별 배치가 갱신).

사용법:
  python scripts/load_short_selling_backfill.py [../reference/backfill_data_response_1.xlsx]
"""
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from sqlalchemy import func, text
from sqlalchemy.dialects.postgresql import insert as pg_insert

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.short_selling import ShortSelling  # noqa: E402

CODE_ROW = 8
ITEM_CODE_ROW = 10
DATA_START_ROW = 15
BATCH_SIZE = 5000

# Item Code -> (short_selling 컬럼, 원단위 변환 배율)
ITEM_SPECS = {
    "S102310": ("short_volume", 1),  # 차입공매도수량 (Shares)
    "S102340": ("short_value", 1_000),  # 차입공매도금액 (Local thou -> 원)
    "S100900": ("total_volume", 1),  # 거래량 (Shares)
    "S101200": ("total_value", 1),  # 거래대금 (Local = 원)
}


def _clean_ticker(code) -> str:
    code = str(code).strip()
    return code[1:] if code.startswith("A") else code


def scan_sheets(wb) -> dict[tuple[str, ...], dict[str, str]]:
    """{청크 티커튜플: {item_code: sheet_name}} — 중복 시트는 자동으로 하나만 남는다."""
    chunks: dict[tuple[str, ...], dict[str, str]] = defaultdict(dict)
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(min_row=CODE_ROW, max_row=ITEM_CODE_ROW, values_only=True)
        row_code = next(it)
        next(it)
        row_item = next(it)
        item_code = row_item[1]
        tickers = tuple(_clean_ticker(c) for c in row_code[1:] if c)
        if item_code not in ITEM_SPECS or not tickers:
            print(f"  건너뜀: {sn} (item_code={item_code!r})")
            continue
        if item_code in chunks[tickers]:
            print(f"  중복 시트 무시: {sn} (item_code={item_code}, 청크 선두={tickers[0]})")
            continue
        chunks[tickers][item_code] = sn
    return chunks


def read_sheet_values(ws, tickers: tuple[str, ...], scale: int) -> dict[tuple[str, object], int]:
    """{(ticker, date): 값} — 결측/비수치는 제외."""
    out: dict[tuple[str, object], int] = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        d = row[0]
        if d is None:
            continue
        d = d.date() if hasattr(d, "date") else d
        for ticker, cell in zip(tickers, row[1:]):
            if cell is None or isinstance(cell, str):
                continue
            out[(ticker, d)] = round(float(cell) * scale)
    return out


def main(path: str):
    print(f"reading {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chunks = scan_sheets(wb)
    print(f"청크 {len(chunks)}개 인식")

    db = SessionLocal()
    instruments_by_ticker = {t: i for t, i in db.query(Instrument.ticker, Instrument.id).all()}

    grand_total = 0
    missing_items_report: dict[str, int] = defaultdict(int)

    for idx, (tickers, item_sheets) in enumerate(chunks.items(), start=1):
        missing = set(ITEM_SPECS) - set(item_sheets)
        for mi in missing:
            missing_items_report[mi] += len(tickers)
        print(f"[청크 {idx}/{len(chunks)}] {len(tickers)}종목, 항목 {sorted(item_sheets)}"
              + (f" (누락: {sorted(missing)})" if missing else ""))

        merged: dict[tuple[str, object], dict[str, int]] = defaultdict(dict)
        for item_code, sheet_name in item_sheets.items():
            column, scale = ITEM_SPECS[item_code]
            values = read_sheet_values(wb[sheet_name], tickers, scale)
            for key, v in values.items():
                merged[key][column] = v

        rows = []
        for (ticker, d), fields in merged.items():
            iid = instruments_by_ticker.get(ticker)
            if iid is None:
                continue
            sv, tv = fields.get("short_volume"), fields.get("total_volume")
            sval, tval = fields.get("short_value"), fields.get("total_value")
            rows.append(
                dict(
                    instrument_id=iid,
                    date=d,
                    short_volume=sv,
                    total_volume=tv,
                    volume_ratio=round(sv / tv * 100, 4) if sv is not None and tv else None,
                    short_value=sval,
                    total_value=tval,
                    value_ratio=round(sval / tval * 100, 4) if sval is not None and tval else None,
                )
            )

        for i in range(0, len(rows), BATCH_SIZE):
            stmt = pg_insert(ShortSelling).values(rows[i : i + BATCH_SIZE])
            # 기존 값은 보존하고 NULL인 컬럼만 채운다(COALESCE) — 겹치는 구간의 pykrx
            # 적재분을 덮어쓰지 않으면서, 나중에 누락분만 따로 받아 다시 돌릴 수 있게 한다.
            stmt = stmt.on_conflict_do_update(
                index_elements=["instrument_id", "date"],
                set_={
                    c: func.coalesce(getattr(ShortSelling, c), getattr(stmt.excluded, c))
                    for c in ("short_volume", "total_volume", "volume_ratio",
                              "short_value", "total_value", "value_ratio")
                },
            )
            db.execute(stmt)
        db.commit()

        grand_total += len(rows)
        print(f"  {len(rows)}행 처리 (누적 {grand_total:,})")

    wb.close()

    # 분모(total_volume/total_value)가 뒤늦게 채워진 행은 비중이 NULL로 남아 있으므로
    # 여기서 채운다. 이미 값이 있는 비중은 건드리지 않는다.
    print("\n비중(ratio) 보정 중...")
    r1 = db.execute(text("""
        update short_selling set volume_ratio = round((short_volume::numeric / total_volume) * 100, 4)
         where volume_ratio is null and short_volume is not null and total_volume > 0"""))
    r2 = db.execute(text("""
        update short_selling set value_ratio = round((short_value::numeric / total_value) * 100, 4)
         where value_ratio is null and short_value is not null and total_value > 0"""))
    db.commit()
    print(f"  volume_ratio {r1.rowcount:,}행 / value_ratio {r2.rowcount:,}행 채움")

    db.close()

    if missing_items_report:
        print("\n항목 누락 요약(해당 종목수):")
        for item_code, n in missing_items_report.items():
            print(f"  {item_code}({ITEM_SPECS[item_code][0]}): {n}종목 분량 시트 없음")
    print("완료.")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "../reference/backfill_data_response_1.xlsx"
    main(target)
