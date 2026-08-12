"""DataGuide "가로형" 응답(시트 1장에 종목×항목이 열로 펼쳐진 형식)에서 가격을 적재한다.

load_full_backfill.py 가 다루는 형식(항목별 시트 6장)과 레이아웃이 달라 별도로 둔다.
개별 종목의 수정주가를 재요청해서 받아올 때 이 형식으로 온다.

    8행  Code       A052670 A052670 ... A053580 ...
    9행  Name       제일바이오 제일바이오 ...
   10행  Item Code  S100300 S100310 ...
   14행  (항목명)    수정주가 수정시가 수정고가 수정저가 종가
   15행~ 날짜별 값

항목명 → 컬럼 매핑은 load_full_backfill.py 와 동일하되, 이 형식은 "수정종가"가 아니라
"수정주가"로 온다. 적재 후 영향받은 종목의 월봉·배당조정 수정종가를 다시 계산한다.

사용법: python scripts/load_wisefn_wide_prices.py ../reference/backfill.xlsx [--dry-run]
"""
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.market_holiday import MarketHoliday  # noqa: E402
from app.models.price import Price  # noqa: E402
from app.services.derived_prices import recompute_dividend_adjusted, recompute_monthly_bar  # noqa: E402

HEADER_START_ROW = 8  # Code 행
LABEL_OFFSET = 6  # Code 행으로부터 항목명 행까지의 거리
FIELD_MAP = {"수정주가": "close", "수정시가": "open", "수정고가": "high", "수정저가": "low", "종가": "raw_close"}


def read_wide(path: str) -> dict[str, dict]:
    """{ticker: {date: {컬럼: 값}}} 로 읽는다."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=HEADER_START_ROW, values_only=True))
    wb.close()

    codes, labels, data = rows[0], rows[LABEL_OFFSET], rows[LABEL_OFFSET + 1 :]

    # (티커 -> {컬럼명: 열번호})
    columns: dict[str, dict[str, int]] = {}
    for j in range(1, len(codes)):
        code, label = codes[j], labels[j]
        if not code or label not in FIELD_MAP:
            continue
        ticker = str(code).strip().lstrip("A")
        columns.setdefault(ticker, {})[FIELD_MAP[label]] = j

    out: dict[str, dict] = {t: {} for t in columns}
    for row in data:
        d = row[0]
        if d is None:
            continue
        d = d.date() if isinstance(d, datetime) else d
        for ticker, colmap in columns.items():
            values = {c: float(row[j]) for c, j in colmap.items() if row[j] is not None}
            if values.get("close"):  # 수정주가 없는 날은 거래 없는 날
                out[ticker][d] = values
    return out


def main(path: str, dry_run: bool = False):
    db = SessionLocal()
    by_ticker = {t: i for t, i in db.query(Instrument.ticker, Instrument.id).all()}
    holidays = {r[0] for r in db.query(MarketHoliday.date).all()}
    names = {t: n for t, n in db.query(Instrument.ticker, Instrument.name).all()}

    parsed = read_wide(path)
    print(f"{path}: {len(parsed)}종목")

    touched: dict[int, set] = {}
    total_rows = 0
    for ticker, by_date in sorted(parsed.items()):
        iid = by_ticker.get(ticker)
        if iid is None:
            print(f"  {ticker}: instruments에 없음 — 건너뜀")
            continue

        existing = {
            r.date: r
            for r in db.query(Price).filter(Price.instrument_id == iid, Price.period == "D").all()
        }
        rows, changed = [], 0
        for d, values in sorted(by_date.items()):
            if d.weekday() >= 5 or d in holidays:
                continue
            prev = existing.get(d)
            if prev is not None and all(
                prev_v is not None and abs(float(prev_v) - values[c]) < 0.5
                for c, prev_v in ((c, getattr(prev, c)) for c in values)
            ):
                rows.append({"instrument_id": iid, "date": d, "period": "D", **values})
                continue
            changed += 1
            rows.append({"instrument_id": iid, "date": d, "period": "D", **values})

        print(f"  {ticker} {names.get(ticker, '')[:14]:<16} {len(rows):>5}행 (변경 {changed}행)")
        total_rows += len(rows)
        if changed and not dry_run:
            for i in range(0, len(rows), 5000):
                stmt = pg_insert(Price).values(rows[i : i + 5000])
                stmt = stmt.on_conflict_do_update(
                    index_elements=["instrument_id", "date", "period"],
                    set_={c: getattr(stmt.excluded, c) for c in ("open", "high", "low", "close", "raw_close")},
                )
                db.execute(stmt)
            db.commit()
            touched[iid] = {(d.year, d.month) for d in by_date}

    if dry_run:
        print(f"\n[dry-run] {total_rows}행 — 적재하지 않음")
        db.close()
        return

    for iid, months in touched.items():
        for year, month in sorted(months):
            recompute_monthly_bar(db, iid, year, month)
        # force_full 필수 — 기본값(증분)은 "마지막 저장일 이후 새 날짜"만 계산해서,
        # 과거 close가 소급 수정된 이번 같은 경우엔 아무것도 안 고친다.
        recompute_dividend_adjusted(db, iid, force_full=True)
        db.commit()
        print(f"  파생 재계산 완료: instrument_id={iid} (월봉 {len(months)}개월)")

    db.close()
    print(f"\n완료 — {len(touched)}종목 갱신")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__)
        sys.exit(1)
    main(args[0], dry_run="--dry-run" in sys.argv)
