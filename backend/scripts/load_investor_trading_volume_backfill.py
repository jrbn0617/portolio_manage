"""1회성 백필 응답(reference/backfill_data_response_2.xlsx)의 투자자별 순매수*수량*을
investor_trading.net_volume 컬럼에 채운다. 정기 실행용이 아님.

load_investor_trading_backfill.py가 순매수*대금*(net_value)을 이미 적재해둔 상태라,
같은 (instrument_id, date, investor_type) 행에 수량만 덧붙이는 성격이다.

주의 1 — 시트명을 믿으면 안 된다: 응답 파일 시트명이 "U110320_*"(대금 백필 때 쓴 이름)
그대로 남아있고 실제 Item Code(U1x0310 계열)와 무관하다. 행10(Item Code)을 읽어 판별한다.

주의 2 — 단위: 원본이 "Shares thou"(천주)라 x1000 해서 주 단위로 저장한다. 즉 이 백필의
수량은 1,000주 단위로 반올림된 값이다.

주의 3 — pykrx 데이터는 덮어쓰지 않는다: daily_update.py가 2026-08-04부터 pykrx로
매수/매도 분해까지 정확한 주 단위로 적재 중이다. net_volume이 이미 있는 행은 건드리지
않고(NULL인 행만 채움), 행 자체가 없으면 새로 만든다.

사용법:
  python scripts/load_investor_trading_volume_backfill.py [../reference/backfill_data_response_2.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.investor_trading import InvestorTrading  # noqa: E402

CODE_ROW = 8
ITEM_CODE_ROW = 10
DATA_START_ROW = 15
BATCH_SIZE = 5000
SCALE = 1_000  # Shares thou(천주) -> 주

# Item Code -> investor_type (daily_update.py의 INVESTOR_TYPES 표기와 동일)
ITEM_CODE_TO_INVESTOR_TYPE = {
    "U110310": "기관합계",
    "U130310": "외국인",
    "U140310": "개인",
}


def _clean_ticker(code) -> str:
    code = str(code).strip()
    return code[1:] if code.startswith("A") else code


def main(path: str):
    print(f"reading {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    db = SessionLocal()
    instruments_by_ticker = {t: i for t, i in db.query(Instrument.ticker, Instrument.id).all()}

    grand_total = 0
    seen_chunks: set[tuple[str, tuple[str, ...]]] = set()

    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(min_row=CODE_ROW, max_row=ITEM_CODE_ROW, values_only=True)
        row_code = next(it)
        next(it)
        item_code = next(it)[1]

        investor_type = ITEM_CODE_TO_INVESTOR_TYPE.get(item_code)
        if investor_type is None:
            print(f"  건너뜀: {sn} (item_code={item_code!r})")
            continue
        tickers = tuple(_clean_ticker(c) for c in row_code[1:] if c)
        if (item_code, tickers) in seen_chunks:
            print(f"  중복 시트 무시: {sn} ({item_code}, 선두={tickers[0]})")
            continue
        seen_chunks.add((item_code, tickers))

        rows = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            d = row[0]
            if d is None:
                continue
            d = d.date() if hasattr(d, "date") else d
            for ticker, cell in zip(tickers, row[1:]):
                if cell is None or isinstance(cell, str):
                    continue
                iid = instruments_by_ticker.get(ticker)
                if iid is None:
                    continue
                rows.append(
                    dict(
                        instrument_id=iid,
                        date=d,
                        investor_type=investor_type,
                        net_volume=round(float(cell) * SCALE),
                    )
                )

        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            stmt = pg_insert(InvestorTrading).values(batch)
            # 이미 net_volume이 있는 행(= pykrx 적재분)은 보존하고, NULL인 행만 채운다.
            stmt = stmt.on_conflict_do_update(
                index_elements=["instrument_id", "date", "investor_type"],
                set_={"net_volume": stmt.excluded.net_volume},
                where=InvestorTrading.net_volume.is_(None),
            )
            db.execute(stmt)
        db.commit()

        grand_total += len(rows)
        print(f"  {sn} [{item_code} {investor_type}] {len(rows):,}행 (누적 {grand_total:,})")

    wb.close()
    db.close()
    print("완료.")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "../reference/backfill_data_response_2.xlsx"
    main(target)
