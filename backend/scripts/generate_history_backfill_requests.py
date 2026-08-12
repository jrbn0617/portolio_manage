"""백테스트 구간을 2015년까지 늘리기 위한 DataGuide 요청 양식을 항목군별로 만든다.

표본이 6.6년(79회 리밸런싱)뿐이라 파라미터 8종 이상을 탐색한 것 대비 자유도가 부족하다.
오버피팅을 실제로 줄이는 방법은 표본을 늘리는 것뿐이고, 그러려면 과거 데이터가 필요하다.
지수 편입 이력은 KRX에서 2015-06-30(코스피200)·2015-07-31(코스닥150)까지 확보했으므로
(backfill_index_memberships_2015.py) 이제 가격·팩터·수급·공매도를 채우면 된다.

**파일을 4개로 나누는 이유**: 한 파일에 다 넣으면 시트가 100장을 넘어가 WISEfn Add-in
새로고침이 사실상 불가능하다. 항목군별로 쪼개면 주기(D/M)도 파일 단위로 일관되게 잡힌다.

  1_가격      (일간) 수정주가·시가/고가/저가·종가·시가총액        -> prices
  2_월간팩터  (월간) 상장주식수·유동주식비율·EBITDA 3종            -> monthly_fundamentals
  3_수급      (일간) 기관합계/외국인/개인 순매수 수량·대금         -> investor_trading
  4_공매도    (일간) 차입공매도 수량·금액 + 거래량·거래대금(분모)  -> short_selling

거래량(S100900)·거래대금(S101200)을 공매도 파일에 넣은 것은 load_short_selling_backfill.py가
그 둘을 short_selling.total_volume/total_value로 적재하기 때문이다(공매도 비중의 분모).

시작일은 2014-01-01로 잡는다. 첫 리밸런싱 가능 시점이 2015-07-31(두 지수 편입 이력이
모두 갖춰지는 때)인데, 12개월 모멘텀 + 최근 1개월 제외 = 13개월치 월말 데이터가 그 앞에
있어야 하므로 최소 2014-07이 필요하고, 여유를 둬서 2014-01로 한다.

유니버스 기본값은 **코스피200·코스닥150 편입 이력이 한 번이라도 있는 종목 전체(727종목)**다.
지금은 상장폐지·합병된 종목이 포함되며(2015년 구 삼성물산·현대증권·LG생명과학 등 11종목은
우리 DB에 시세가 아예 없다), **이들을 빼면 생존편향이 생겨 표본을 늘리는 의미가 사라진다.**
`--universe all`을 주면 DB의 전체 보통주(약 3,012종목)로 넓힐 수 있으나 요청량이 4배가 되고,
코스피전체/코스닥전체 유니버스를 과거로 돌리려면 KOSPI/KOSDAQ 시장구분 스냅샷도 따로
받아야 한다(현재 2018-12-31부터).

사용법:
  python scripts/generate_history_backfill_requests.py
  python scripts/generate_history_backfill_requests.py --from 2014-01-01 --universe all
"""
import argparse
import datetime
import sys
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import text  # noqa: E402

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.services.instrument_rules import is_common_stock  # noqa: E402

# reference/ 는 gitignore 대상이라 PC에 따라 원본 템플릿이 없을 수 있다. 그 경우 과거에
# 생성한 요청 파일의 첫 시트를 템플릿으로 재사용한다 — Add-in 구조(A1 Refresh, A3 Time
# Series, A4 Frequency, A5/A6 Period, 8~12행 라벨, A14 D A T E)가 그대로 들어 있다.
# 단 그 시트는 500종목이 이미 채워져 있으므로 **재사용 전에 반드시 비워야 한다**
# (마지막 청크가 500개 미만이면 이전 종목코드가 남아 잘못된 요청이 나간다).
TEMPLATE_CANDIDATES = [
    (REPO_DIR / "reference" / "monthly_data_template.xlsx", "sample"),
    (REPO_DIR / "reference" / "backfill_data_request.xlsx", None),  # None = 첫 시트
]
OUTPUT_DIR = REPO_DIR / "reference"
CHUNK_SIZE = 500

DEFAULT_FROM = datetime.date(2014, 1, 1)
FREQUENCY_CELL, PERIOD_FROM_CELL, PERIOD_TO_CELL = "B4", "B5", "B6"
CODE_ROW, NAME_ROW, ITEM_CODE_ROW, UNIT_ROW, BASE_DATE_ROW = 8, 9, 10, 11, 12
DATE_LIST_ROWS = range(15, 21)  # 템플릿 잔여 날짜 — Add-in이 채우므로 비운다

# 항목별로 **우리 DB에 이미 있는 데이터의 시작일**. Period(To)는 이 날의 전날로 잡아서
# 이미 가진 구간을 다시 받지 않는다(2026-08-12 실측):
#   prices(close/open/high/low/raw_close/market_cap/volume)  2018-12-28
#   monthly_fundamentals: 상장주식수·EBITDA 3종               2018-12-28
#                         유동주식비율                          2019-12-30  <- 이것만 1년 늦다
#   investor_trading / short_selling                          2018-12-28
EXISTING_START_DEFAULT = datetime.date(2018, 12, 28)
EXISTING_START = {"S102060": datetime.date(2019, 12, 30)}  # 유동주식비율


def period_to_for(item_code: str) -> datetime.date:
    return EXISTING_START.get(item_code, EXISTING_START_DEFAULT) - datetime.timedelta(days=1)


# (파일 접미어, Frequency, [(시트명 접두어, Item Code, Unit, Base Date), ...])
FILE_SPECS = [
    (
        "1_가격",
        "D",
        [
            ("수정주가", "S100300", "Local/Shares", ""),
            ("수정시가", "S100310", "Local/Shares", ""),
            ("수정고가", "S100320", "Local/Shares", ""),
            ("수정저가", "S100330", "Local/Shares", ""),
            ("종가", "S100100", "Local/Shares", ""),
            ("시가총액", "S102100", "Local", ""),
        ],
    ),
    (
        "2_월간팩터",
        "M",
        [
            ("상장주식수", "S101500", "Shares", ""),
            ("유동주식비율", "S102060", "%", ""),
            ("EBITDA(TTM)", "M123005.M", "Local thou", "NR.FY1"),
            ("EBITDA(Fwd12M)", "E123060.M", "Local thou", ""),
            ("EV EBITDA(Fwd12M)", "E331060.M", "X", ""),
        ],
    ),
    (
        "3_수급",
        "D",
        [
            ("기관합계_수량", "U110310", "Shares thou", ""),
            ("외국인_수량", "U130310", "Shares thou", ""),
            ("개인_수량", "U140310", "Shares thou", ""),
            ("기관합계_대금", "U110320", "Local mn", ""),
            ("외국인_대금", "U130320", "Local mn", ""),
            ("개인_대금", "U140320", "Local mn", ""),
        ],
    ),
    (
        "4_공매도",
        "D",
        [
            ("차입공매도수량", "S102310", "Shares", "CPD"),
            ("차입공매도금액", "S102340", "Local thou", "CPD"),
            ("거래량", "S100900", "Shares", ""),
            ("거래대금", "S101200", "Local", ""),
        ],
    ),
]

UNIVERSE_SQL = {
    # 코스피200·코스닥150에 한 번이라도 편입된 종목 (상폐·합병분 포함)
    "index": """
        select distinct i.ticker, i.name
        from instruments i
        join index_memberships m on m.instrument_id = i.id
        where m.index_name in ('KOSPI200','KOSDAQ150')
          and coalesce(i.asset_type,'') <> 'index'
        order by 1
    """,
    # DB가 아는 전체 종목 (일별 시세가 있거나 어떤 지수에든 편입 이력이 있는 종목)
    "all": """
        select distinct i.ticker, i.name
        from instruments i
        where coalesce(i.asset_type,'') <> 'index'
          and (exists (select 1 from prices p where p.instrument_id = i.id and p.period='D')
               or exists (select 1 from index_memberships m where m.instrument_id = i.id))
        order by 1
    """,
}


def fetch_universe(db, scope: str) -> list[tuple[str, str]]:
    rows = db.execute(text(UNIVERSE_SQL[scope])).fetchall()
    # 우선주는 DB에서 이미 제거했지만, 남아 있어도 요청에는 넣지 않는다(종목코드 끝자리 규칙).
    return [(r.ticker, r.name or r.ticker) for r in rows if is_common_stock(r.ticker)]


def chunked(items: list, size: int) -> list[list]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def resolve_template() -> tuple[Path, str]:
    for path, sheet in TEMPLATE_CANDIDATES:
        if path.exists():
            wb = openpyxl.load_workbook(path)
            name = sheet if sheet and sheet in wb.sheetnames else wb.sheetnames[0]
            wb.close()
            return path, name
    raise SystemExit(
        "템플릿을 찾지 못했습니다. reference/monthly_data_template.xlsx 또는 "
        "과거 요청 파일(backfill_data_request.xlsx)이 필요합니다."
    )


def clear_sheet(ws):
    """템플릿에 남아있는 종목·날짜·데이터를 전부 비운다."""
    for row in list(range(CODE_ROW, BASE_DATE_ROW + 1)) + list(DATE_LIST_ROWS):
        for col in range(2, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    for row in DATE_LIST_ROWS:
        ws.cell(row=row, column=1).value = None


def fill_sheet(ws, chunk, frequency: str, period_from: datetime.date, period_to: datetime.date, item_code: str, unit: str, base_date: str):
    clear_sheet(ws)
    ws[FREQUENCY_CELL] = frequency
    ws[PERIOD_FROM_CELL] = period_from
    ws[PERIOD_TO_CELL] = period_to

    for offset, (ticker, name) in enumerate(chunk):
        col = 2 + offset  # B열부터 가로로
        ws.cell(row=CODE_ROW, column=col).value = f"A{ticker}"
        ws.cell(row=NAME_ROW, column=col).value = name
        ws.cell(row=ITEM_CODE_ROW, column=col).value = item_code
        ws.cell(row=UNIT_ROW, column=col).value = unit
        ws.cell(row=BASE_DATE_ROW, column=col).value = base_date


def build_file(
    suffix: str, frequency: str, specs: list, universe: list, period_from: datetime.date, template: tuple[Path, str]
) -> Path:
    tpl_path, tpl_sheet = template
    wb = openpyxl.load_workbook(tpl_path)
    sample = wb[tpl_sheet]
    chunks = chunked(universe, CHUNK_SIZE)

    keep = set()
    for prefix, item_code, unit, base_date in specs:
        for i, chunk in enumerate(chunks, start=1):
            ws = wb.copy_worksheet(sample)
            ws.title = f"{prefix}{i}" if len(chunks) > 1 else prefix
            fill_sheet(ws, chunk, frequency, period_from, period_to_for(item_code), item_code, unit, base_date)
            keep.add(ws.title)

    for name in [n for n in wb.sheetnames if n not in keep]:
        del wb[name]
    out = OUTPUT_DIR / f"백필요청_{suffix}_{period_from:%Y%m}.xlsx"
    wb.save(out)
    tos = sorted({period_to_for(c) for _, c, _, _ in specs})
    print(
        f"  {out.name}  주기={frequency}  항목 {len(specs)}개 x 청크 {len(chunks)}개 = 시트 {len(wb.sheetnames)}장"
        f"  Period(To)={', '.join(str(t) for t in tos)}"
    )
    return out


def main(period_from: datetime.date, scope: str):
    db = SessionLocal()
    universe = fetch_universe(db, scope)
    # distinct 없이 세면 스냅샷 수만큼 중복 집계된다(11종목이 44로 나왔던 버그).
    nopx = db.execute(
        text(
            """select count(distinct i.id) from instruments i
               join index_memberships m on m.instrument_id=i.id
               where m.index_name in ('KOSPI200','KOSDAQ150')
                 and not exists (select 1 from prices p where p.instrument_id=i.id and p.period='D')"""
        )
    ).scalar()
    print(f"유니버스({scope}): {len(universe)}종목 — 우선주·스팩 제외")
    print(f"  그중 DB에 시세가 전혀 없는 종목: {nopx}개 (상폐·합병분 — 이걸 빼면 생존편향)")
    print(f"시작일: {period_from}\n")

    template = resolve_template()
    print(f"템플릿: {template[0].name} / 시트 '{template[1]}'\n")
    for suffix, frequency, specs in FILE_SPECS:
        build_file(suffix, frequency, specs, universe, period_from, template)
    db.close()


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--from", dest="period_from", type=datetime.date.fromisoformat, default=DEFAULT_FROM)
    p.add_argument("--universe", choices=list(UNIVERSE_SQL), default="index")
    a = p.parse_args()
    main(a.period_from, a.universe)
