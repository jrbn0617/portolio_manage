"""EBITDA PEG 스크리닝 + 모멘텀 백테스트 결과를
reference/(별첨 2) 백테스팅자료_회사명(예시).xlsx 포맷으로 출력한다.
(export_backtest_excel.py와 동일한 3시트 구성 — 순수 모멘텀이 아니라
스크리닝을 더한 버전이라 별도 스크립트로 분리했다.)

시트 3개: 작성가이드(원본 문구 그대로), 시계열(전략 NAV vs 벤치마크, 공식
기간 첫 거래일 기준 100 리베이스), 리밸런싱발생내역(리밸런싱별 종목 비중 매트릭스).

사전 조건: run_screened_momentum_backtest.py로 같은 기간을 먼저 실행해 미분류
시계열단절 경고가 0건인지 확인해야 한다 — 있으면 resolve_backtest_event.py로 먼저 해결.

사용법: python scripts/export_screened_backtest_excel.py [--start 2020-01-01] [--end 2026-04-30]
        [--out "../reference/(별첨 2) 백테스팅자료_코스닥150 EBITDAPEG스크리닝모멘텀.xlsx"]
"""
import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db.base import Base  # noqa: F401
from app.db.session import SessionLocal
from app.models.instrument import Instrument
from app.models.price import Price
from app.services.backtest_service import (
    BacktestConfig,
    compute_free_float_weights,
    compute_inverse_momentum_weights,
    compute_inverse_vol_weights,
    compute_momentum_weights,
    run_momentum_backtest,
)
from app.services.factor_screen_service import ScreenConfig, screen_by_ebitda_peg
from functools import partial

GUIDE_LINES = [
    "* 백테스팅 기간은 2020.1.1 ~ 2026.4.30 로 맞춰주시기 바랍니다.",
    '* 리밸런싱발생내역의 자산종류명은 알고리즘설명서의 "편입자산 종류 및 특징"의 자산종류명과 같아야 합니다.',
    "* 시계열 시트에 알고리즘의 시계열과 BM의 시계열을 함께 넣어주시기 바랍니다.",
    "* 날짜는 영업일 기준으로 작성해주세요.",
]

INDEX_LABEL_MAP = {"KOSPI": "코스피전체", "KOSPI200": "코스피200", "KOSDAQ": "코스닥전체", "KOSDAQ150": "코스닥150"}
BM_LABEL_MAP = {
    "KOSPI": "KOSPI(BM)",
    "KOSPI200": "KOSPI200(BM)",
    "KOSDAQ": "KOSDAQ(BM)",
    "KOSDAQ150": "KOSDAQ150TR(BM)",
}
ASSET_TYPE_LABEL = "코스닥150 EBITDA PEG스크리닝 모멘텀"  # 기본값(레거시 호출 호환용), main()에서 --index 기준으로 재계산
BENCHMARK_TICKER = "KOSDAQ150"
BENCHMARK_LABEL = "KOSDAQ150TR(BM)"
REBALANCE_REASON = "정기리밸런싱"
PCT_FMT = "0.0%"


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--index", choices=list(INDEX_LABEL_MAP), default="KOSDAQ150")
    p.add_argument("--start", type=date.fromisoformat, default=date(2020, 1, 1))
    p.add_argument("--end", type=date.fromisoformat, default=date(2026, 4, 30))
    p.add_argument("--top-n", type=int, default=10)
    p.add_argument("--max-per-sector", type=int, default=None)
    p.add_argument("--screen-top-pct", type=float, default=0.5)
    p.add_argument("--min-sector-size", type=int, default=5)
    p.add_argument("--ttm-lag-days", type=int, default=90)
    p.add_argument("--consensus-lag-days", type=int, default=0)
    p.add_argument("--peg-min", type=float, default=0.0)
    p.add_argument(
        "--weighting", choices=["equal", "free_float", "risk_budget", "momentum", "inverse_momentum"], default="equal"
    )
    p.add_argument("--vol-lookback-days", type=int, default=63, help="risk_budget 변동성 추정 lookback 거래일수")
    p.add_argument("--smoothing-multiplier", type=float, default=2.0, help="inverse_momentum 스무딩 상수(평균의 N배)")
    p.add_argument("--max-weight", type=float, default=None, help="종목당 최대 편입비중(예: 0.3 = 30%)")
    p.add_argument("--out", default=None)
    return p.parse_args()


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("작성가이드")
    for i, line in enumerate(GUIDE_LINES, start=3):
        ws.cell(row=i, column=2, value=line)


def build_timeseries_sheet(
    wb: Workbook, dates: list[date], mp_values: list[float], bm_values: list[float], bm_label: str = BENCHMARK_LABEL
):
    ws = wb.create_sheet("시계열")
    ws.cell(row=1, column=2, value="mp")
    ws.cell(row=1, column=3, value=bm_label)
    for i, (d, mp, bm) in enumerate(zip(dates, mp_values, bm_values), start=2):
        ws.cell(row=i, column=1, value=d).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=mp)
        ws.cell(row=i, column=3, value=bm)


def build_rebalance_sheet(
    wb: Workbook, rebalances: list[dict], instruments_by_ticker: dict[str, Instrument], asset_label: str = ASSET_TYPE_LABEL
):
    ws = wb.create_sheet("리밸런싱발생내역")

    tickers: list[str] = []
    seen = set()
    for r in rebalances:
        for t in r["holdings"]:
            if t not in seen:
                seen.add(t)
                tickers.append(t)

    n = len(tickers)
    reason_col = n + 2  # A=1(라벨), B..=종목, 마지막=리밸런싱 사유

    ws.cell(row=1, column=1, value="자산종류")
    ws.cell(row=1, column=2, value=asset_label)
    if n > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws.cell(row=1, column=reason_col, value="리밸런싱 사유")
    ws.merge_cells(start_row=1, start_column=reason_col, end_row=3, end_column=reason_col)
    ws.cell(row=1, column=reason_col).alignment = Alignment(vertical="center")

    ws.cell(row=2, column=1, value="종목명")
    ws.cell(row=3, column=1, value="업종")
    for j, ticker in enumerate(tickers, start=2):
        inst = instruments_by_ticker[ticker]
        ws.cell(row=2, column=j, value=inst.name)
        ws.cell(row=3, column=j, value=inst.industry or inst.krx_sector)

    col_of_ticker = {t: j for j, t in enumerate(tickers, start=2)}
    for i, r in enumerate(rebalances, start=4):
        ws.cell(row=i, column=1, value=r["date"]).number_format = "yyyy-mm-dd"
        for t in r["holdings"]:
            cell = ws.cell(row=i, column=col_of_ticker[t], value=r["weights"][t])
            cell.number_format = PCT_FMT
        ws.cell(row=i, column=reason_col, value=REBALANCE_REASON)

    ws.column_dimensions[get_column_letter(1)].width = 12


def main():
    args = parse_args()
    benchmark_ticker = args.index
    benchmark_label = BM_LABEL_MAP[args.index]
    index_label = INDEX_LABEL_MAP[args.index]

    config = BacktestConfig(
        index_name=args.index,
        top_n=args.top_n,
        max_per_sector=args.max_per_sector,
        start_date=args.start,
        end_date=args.end,
    )
    screen_config = ScreenConfig(
        top_pct=args.screen_top_pct,
        min_sector_size=args.min_sector_size,
        ttm_lag_days=args.ttm_lag_days,
        consensus_lag_days=args.consensus_lag_days,
        peg_min=args.peg_min,
    )

    db = SessionLocal()
    screen_fn = partial(screen_by_ebitda_peg, config=screen_config, warn=print, info=lambda m: None)
    if args.weighting == "free_float":
        weight_fn = partial(compute_free_float_weights, max_weight=args.max_weight)
    elif args.weighting == "risk_budget":
        weight_fn = partial(
            compute_inverse_vol_weights, lookback_days=args.vol_lookback_days, max_weight=args.max_weight
        )
    elif args.weighting == "momentum":
        weight_fn = partial(compute_momentum_weights, max_weight=args.max_weight)
    elif args.weighting == "inverse_momentum":
        weight_fn = partial(
            compute_inverse_momentum_weights, smoothing_multiplier=args.smoothing_multiplier, max_weight=args.max_weight
        )
    else:
        weight_fn = None
    result = run_momentum_backtest(
        db, config, on_warning=print, on_info=lambda m: None, screen_fn=screen_fn, weight_fn=weight_fn
    )

    if result.warnings:
        print(f"\n미분류 시계열단절 {len(result.warnings)}건이 남아있습니다. resolve_backtest_event.py로 먼저 해결하세요:")
        for w in result.warnings:
            print(f"  - {w}")
        db.close()
        sys.exit(1)

    # 공식 기간 첫 거래일(예: 2020-01-02) 기준으로 100 리베이스, 그 이전(초기 형성일)은 시계열 시트에서 제외
    official_dates = [d for d, _ in result.nav_series if d >= args.start]
    anchor_date = official_dates[0]
    nav_by_date = dict(result.nav_series)
    anchor_nav = nav_by_date[anchor_date]

    bench_rows = (
        db.query(Price.date, Price.close)
        .join(Instrument, Instrument.id == Price.instrument_id)
        .filter(Instrument.ticker == benchmark_ticker, Price.period == "D", Price.date.in_(official_dates))
        .all()
    )
    bench_by_date = {r.date: float(r.close) for r in bench_rows}
    missing_bench = [d for d in official_dates if d not in bench_by_date]
    if missing_bench:
        print(f"경고: 벤치마크에 없는 날짜 {len(missing_bench)}건 (예: {missing_bench[:5]})")
    anchor_bench = bench_by_date[anchor_date]

    mp_values = [nav_by_date[d] / anchor_nav * 100 for d in official_dates]
    bm_values = [bench_by_date[d] / anchor_bench * 100 for d in official_dates]

    all_tickers = {t for r in result.rebalances for t in r["holdings"]}
    instruments_by_ticker = {
        i.ticker: i for i in db.query(Instrument).filter(Instrument.ticker.in_(all_tickers)).all()
    }

    weighting_suffix = {
        "free_float": "(유동시총가중)",
        "risk_budget": "(리스크버짓-역변동성가중)",
        "momentum": "(모멘텀가중)",
        "inverse_momentum": f"(역모멘텀가중-{args.smoothing_multiplier:g}배스무딩)",
        "equal": "(동일가중)",
    }[args.weighting]
    sector_suffix = f"_섹터당{args.max_per_sector}개이하" if args.max_per_sector is not None else ""
    cap_suffix = f"_종목당최대{args.max_weight:.0%}" if args.max_weight is not None else ""
    asset_label = f"{index_label} EBITDA PEG스크리닝 모멘텀" + weighting_suffix + sector_suffix + cap_suffix

    wb = Workbook()
    wb.remove(wb.active)
    build_guide_sheet(wb)
    build_timeseries_sheet(wb, official_dates, mp_values, bm_values, bm_label=benchmark_label)
    build_rebalance_sheet(wb, result.rebalances, instruments_by_ticker, asset_label=asset_label)

    default_out = (
        f"../reference/(별첨 2) 백테스팅자료_{index_label} EBITDAPEG스크리닝모멘텀_top{args.top_n}"
        f"{sector_suffix}{cap_suffix}{weighting_suffix}.xlsx"
    )
    out_path = Path(args.out or default_out)
    wb.save(out_path)
    db.close()

    print(f"\n저장 완료: {out_path}")
    print(f"  시계열: {len(official_dates)}행 ({official_dates[0]} ~ {official_dates[-1]})")
    print(f"  리밸런싱: {len(result.rebalances)}회, 누적종목수: {len(all_tickers)}개")


if __name__ == "__main__":
    main()
