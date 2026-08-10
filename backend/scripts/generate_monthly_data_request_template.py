"""WISEfn/DataGuide에 매월 수동으로 요청하는 4개 팩터(유동주식비율/EBITDA(TTM)/
EBITDA(Fwd.12M)/EV EBITDA(Fwd.12M))의 요청 양식을 자동으로 만든다.
(docs/plans/01-data-management.md "월간 데이터 수집 체크리스트" 참고. 상장주식수는
pykrx로 자동 수집하므로 이 양식에는 포함하지 않음 — load_shares_outstanding_pykrx.py)

reference/monthly_data_template.xlsx의 'sample' 시트(WISEfn Add-in 포맷)를 복제해서
채우는 방식으로 동작한다 — 새 시트를 백지에서 만들지 않고, 항상 이 템플릿을 원본으로
복사한다(원본 파일 자체는 절대 수정하지 않음).

시트 레이아웃(행8=Code, 행9=Name, 행10=Item Code, 행11=Unit, 행12=Base Date, 행15부터
날짜별 데이터)은 backend/scripts/load_monthly_fundamentals.py / load_factor_fundamentals.py가
이미 파싱하는 레이아웃과 동일 — 데이터가 채워져 돌아오면 그 로더들로 바로 적재 가능.

월말 날짜는 로컬 DB의 prices(period='D')에서 그 달의 최댓값 날짜를 그대로 사용한다
(pykrx 등 원격 조회 없이 DB에 이미 있는 실제 거래일과 100% 일치).

사용법:
  python scripts/generate_monthly_data_request_template.py                  # 당월 기준 6개월치
  python scripts/generate_monthly_data_request_template.py --month 2026-06  # 2026년 6월 기준
"""
import argparse
import datetime
import re
import sys
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import func  # noqa: E402

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.price import Price  # noqa: E402

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "reference" / "monthly_data_template.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parents[2] / "reference"
SAMPLE_SHEET = "sample"
CHUNK_SIZE = 500
LOOKBACK_MONTHS = 6

# 종목명 끝이 숫자+"우"(+"선주")(+대문자 1글자, 신형우선주 구분용)로 끝나면 우선주로 판단.
# 예: "삼성전자1우", "두산퓨얼셀1우선주", "SK이노베이션1우B". 숫자 없이 "우"로만 끝나는
# 이름(연우, 성우 등)은 실제 보통주라 오탐이므로 숫자 접두를 필수로 둔다.
PREFERRED_NAME_RE = re.compile(r"\d우(선주)?[A-Z]?$")
SPAC_NAME_RE = re.compile(r"스팩")


def _is_common_stock(name: str) -> bool:
    return not PREFERRED_NAME_RE.search(name) and not SPAC_NAME_RE.search(name)

CODE_ROW = 8
NAME_ROW = 9
ITEM_CODE_ROW = 10
UNIT_ROW = 11
BASE_DATE_ROW = 12
DATE_START_ROW = 15
DATE_END_ROW = 20  # DATE_START_ROW + LOOKBACK_MONTHS - 1
PERIOD_FROM_CELL = "B5"

# (시트명 접두어, Item Code, Unit, Base Date)
ITEM_SPECS = [
    ("유동비율", "S102060", "%", ""),
    ("EBITDA(TTM)", "M123005.M", "Local thou", "NR.FY1"),
    ("EBITDA(Fwd.12M)", "E123060.M", "Local thou", ""),
    ("EV EBITDA(Fwd.12M)", "E331060.M", "X", ""),
]


def _shift_month(year: int, month: int, offset: int) -> tuple[int, int]:
    total = year * 12 + (month - 1) + offset
    return total // 12, total % 12 + 1


def _month_bounds(year: int, month: int) -> tuple[datetime.date, datetime.date]:
    start = datetime.date(year, month, 1)
    next_y, next_m = _shift_month(year, month, 1)
    end = datetime.date(next_y, next_m, 1) - datetime.timedelta(days=1)
    return start, end


def last_trading_day_of_month(db, year: int, month: int) -> datetime.date:
    start, end = _month_bounds(year, month)
    result = db.query(func.max(Price.date)).filter(
        Price.period == "D", Price.date >= start, Price.date <= end
    ).scalar()
    if result is None:
        raise RuntimeError(f"{year}-{month:02d}: prices(D) 데이터가 DB에 없습니다.")
    return result


def target_months(target_year: int, target_month: int) -> list[tuple[int, int]]:
    months = [_shift_month(target_year, target_month, -offset) for offset in range(LOOKBACK_MONTHS - 1, -1, -1)]
    return months


def fetch_universe(db) -> list[tuple[str, str]]:
    latest_date = db.query(func.max(Price.date)).filter(Price.period == "D").scalar()
    if latest_date is None:
        raise RuntimeError("prices(D) 데이터가 DB에 없습니다.")
    rows = (
        db.query(Instrument.ticker, Instrument.name)
        .join(Price, Price.instrument_id == Instrument.id)
        .filter(Instrument.asset_type == "stock", Price.period == "D", Price.date == latest_date)
        .distinct()
        .order_by(Instrument.ticker)
        .all()
    )
    return [(t, n) for t, n in rows if _is_common_stock(n)]


def _chunk(items: list, size: int) -> list[list]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def fill_sheet(ws, universe_chunk: list[tuple[str, str]], dates: list[datetime.date], item_code: str, unit: str, base_date: str):
    ws[PERIOD_FROM_CELL] = dates[0]
    for i, d in enumerate(dates):
        ws.cell(row=DATE_START_ROW + i, column=1).value = d

    for col_offset, (ticker, name) in enumerate(universe_chunk):
        col = 2 + col_offset  # B열부터
        ws.cell(row=CODE_ROW, column=col).value = f"A{ticker}"
        ws.cell(row=NAME_ROW, column=col).value = name
        ws.cell(row=ITEM_CODE_ROW, column=col).value = item_code
        ws.cell(row=UNIT_ROW, column=col).value = unit
        ws.cell(row=BASE_DATE_ROW, column=col).value = base_date


def build_workbook(template_path: Path, universe: list[tuple[str, str]], dates: list[datetime.date]) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(template_path)
    sample = wb[SAMPLE_SHEET]

    total_sheets = 0
    for prefix, item_code, unit, base_date in ITEM_SPECS:
        chunks = _chunk(universe, CHUNK_SIZE)
        for i, chunk in enumerate(chunks, start=1):
            ws = wb.copy_worksheet(sample)
            ws.title = f"{prefix}{i}"
            fill_sheet(ws, chunk, dates, item_code, unit, base_date)
            total_sheets += 1
        print(f"  {prefix}: {len(chunks)}개 시트 ({len(universe)}종목)")

    del wb[SAMPLE_SHEET]
    print(f"총 {total_sheets}개 시트 생성")
    return wb


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--month", type=str, default=None, help="대상월 YYYY-MM (기본: 당월)")
    p.add_argument("--template", type=Path, default=TEMPLATE_PATH, help="템플릿 파일 경로")
    return p.parse_args()


def main():
    args = parse_args()
    if args.month is not None:
        target_year, target_month = (int(x) for x in args.month.split("-"))
    else:
        today = datetime.date.today()
        target_year, target_month = today.year, today.month

    db = SessionLocal()

    months = target_months(target_year, target_month)
    dates = [last_trading_day_of_month(db, y, m) for y, m in months]
    print(f"대상 {LOOKBACK_MONTHS}개월 마지막 거래일: {[d.isoformat() for d in dates]}")

    universe = fetch_universe(db)
    print(f"유니버스(현재 상장 주식): {len(universe)}종목")

    wb = build_workbook(args.template, universe, dates)

    out_path = OUTPUT_DIR / f"monthly_data_request_{target_year:04d}-{target_month:02d}.xlsx"
    wb.save(out_path)
    print(f"저장 완료: {out_path}")

    db.close()


if __name__ == "__main__":
    main()
