"""1회성 백필용 WISEfn/DataGuide 요청 양식을 만든다 (정기 수집용
generate_monthly_data_request_template.py와 달리 한 번만 돌리는 스크립트).

대상 항목(Item Code, Unit — 전부 Base Date 없음):
  U110320, U140320, U130320 / Local mn

reference/monthly_data_template.xlsx의 'sample' 시트를 복제해서 채우는 방식은
generate_monthly_data_request_template.py와 동일(행8=Code, 행9=Name, 행10=Item Code,
행11=Unit, 행12=Base Date). 다만 이 백필은:
  - Frequency를 D(일간)으로 설정 (템플릿 기본값 M을 덮어씀)
  - Period(From)만 2018-12-28로 설정하고, 왼쪽 날짜 목록(행15부터)은 채우지 않음
    — WISEfn Add-in이 Period(From)~Period(To) 사이를 새로고침 시 알아서 채움
  - 유니버스는 "현재 상장된 종목"이 아니라 DB에 일별 시세가 한 번이라도 존재했던
    모든 종목(상폐 종목 포함) — prices(period='D')에 1998-12-31 이후 데이터가
    있으면 포함. 실제로 DB의 시세 데이터는 2018-12-28부터 시작하므로 사실상
    "DB에 있는 전체 종목"과 동일.

사용법:
  python scripts/generate_backfill_data_request_template.py
"""
import datetime
import re
import sys
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from app.db.base import Base  # noqa: E402,F401
from app.db.session import SessionLocal  # noqa: E402
from app.models.instrument import Instrument  # noqa: E402
from app.models.price import Price  # noqa: E402

TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "reference" / "monthly_data_template.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parents[2] / "reference" / "backfill_data_request.xlsx"
SAMPLE_SHEET = "sample"
CHUNK_SIZE = 500

UNIVERSE_SINCE = "1998-12-31"
PERIOD_FROM = datetime.date(2018, 12, 28)
FREQUENCY = "D"

# 종목명 끝이 숫자+"우"(+"선주")(+대문자 1글자, 신형우선주 구분용)로 끝나면 우선주로 판단
# (generate_monthly_data_request_template.py와 동일 로직 — 재무제표 항목이라 보통주만 대상).
PREFERRED_NAME_RE = re.compile(r"\d우(선주)?[A-Z]?$")
SPAC_NAME_RE = re.compile(r"스팩")


def _is_common_stock(name: str) -> bool:
    return not PREFERRED_NAME_RE.search(name) and not SPAC_NAME_RE.search(name)


FREQUENCY_CELL = "B4"
PERIOD_FROM_CELL = "B5"
CODE_ROW = 8
NAME_ROW = 9
ITEM_CODE_ROW = 10
UNIT_ROW = 11
BASE_DATE_ROW = 12
DATE_LIST_ROWS = range(15, 21)  # 템플릿에 남아있는 기본 날짜 6개 — 이번엔 안 쓰므로 비움

# (시트명 접두어, Item Code, Unit, Base Date)
ITEM_SPECS = [
    ("U110320", "U110320", "Local mn", ""),
    ("U140320", "U140320", "Local mn", ""),
    ("U130320", "U130320", "Local mn", ""),
]


def fetch_universe(db) -> list[tuple[str, str]]:
    rows = (
        db.query(Instrument.ticker, Instrument.name)
        .join(Price, Price.instrument_id == Instrument.id)
        .filter(Instrument.asset_type == "stock", Price.period == "D", Price.date >= UNIVERSE_SINCE)
        .distinct()
        .order_by(Instrument.ticker)
        .all()
    )
    return [(t, n) for t, n in rows if _is_common_stock(n)]


def _chunk(items: list, size: int) -> list[list]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def fill_sheet(ws, universe_chunk: list[tuple[str, str]], item_code: str, unit: str, base_date: str):
    ws[FREQUENCY_CELL] = FREQUENCY
    ws[PERIOD_FROM_CELL] = PERIOD_FROM
    for row in DATE_LIST_ROWS:
        ws.cell(row=row, column=1).value = None

    for col_offset, (ticker, name) in enumerate(universe_chunk):
        col = 2 + col_offset  # B열부터
        ws.cell(row=CODE_ROW, column=col).value = f"A{ticker}"
        ws.cell(row=NAME_ROW, column=col).value = name
        ws.cell(row=ITEM_CODE_ROW, column=col).value = item_code
        ws.cell(row=UNIT_ROW, column=col).value = unit
        ws.cell(row=BASE_DATE_ROW, column=col).value = base_date


def build_workbook(template_path: Path, universe: list[tuple[str, str]]) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(template_path)
    sample = wb[SAMPLE_SHEET]

    total_sheets = 0
    for prefix, item_code, unit, base_date in ITEM_SPECS:
        chunks = _chunk(universe, CHUNK_SIZE)
        for i, chunk in enumerate(chunks, start=1):
            ws = wb.copy_worksheet(sample)
            ws.title = f"{prefix}_{i}"
            fill_sheet(ws, chunk, item_code, unit, base_date)
            total_sheets += 1
        print(f"  {prefix}: {len(chunks)}개 시트 ({len(universe)}종목)")

    del wb[SAMPLE_SHEET]
    print(f"총 {total_sheets}개 시트 생성")
    return wb


def main():
    db = SessionLocal()

    universe = fetch_universe(db)
    print(f"유니버스(DB 내 전체 보통주, {UNIVERSE_SINCE} 이후 시세 존재): {len(universe)}종목")

    wb = build_workbook(TEMPLATE_PATH, universe)
    wb.save(OUTPUT_PATH)
    print(f"저장 완료: {OUTPUT_PATH}")

    db.close()


if __name__ == "__main__":
    main()
